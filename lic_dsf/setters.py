from __future__ import annotations

from dataclasses import dataclass
from collections.abc import Mapping as MappingABC, Sequence as SequenceABC
from typing import Mapping, Sequence

from .internals import CellValue, EvalContext
from .inputs import DEFAULT_INPUTS


@dataclass(frozen=True, slots=True)
class YearSeriesAssignment:
    years: tuple[int, ...]
    applied: dict[int, str]
    ignored: dict[int, CellValue]


@dataclass(frozen=True, slots=True)
class RangeAssignment:
    shape: tuple[int, int]
    addresses: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class YearRowAssignment:
    years: tuple[int, ...]
    applied: dict[int, tuple[str, ...]]
    ignored: dict[int, CellValue]


def _split_sheet_address(address: str) -> tuple[str, str]:
    if '!' not in address:
        raise ValueError(f"Invalid address: {address}")
    if address.startswith("'"):
        i = 1
        sheet_chars: list[str] = []
        while i < len(address):
            ch = address[i]
            if ch == "'":
                if i + 1 < len(address) and address[i + 1] == "'":
                    sheet_chars.append("'")
                    i += 2
                    continue
                if i + 1 < len(address) and address[i + 1] == "!":
                    sheet = "".join(sheet_chars)
                    a1 = address[i + 2 :]
                    if not a1:
                        raise ValueError(f"Invalid address: {address}")
                    return sheet, a1
                raise ValueError(f"Invalid address: {address}")
            sheet_chars.append(ch)
            i += 1
        raise ValueError(f"Invalid address: {address}")
    sheet, a1 = address.split("!", 1)
    if not sheet or not a1:
        raise ValueError(f"Invalid address: {address}")
    return sheet, a1


def _read_inputs_from_workbook(workbook_path: str) -> dict[str, CellValue]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required to read inputs from a workbook") from exc
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)
    try:
        updates: dict[str, CellValue] = {}
        ws_cache: object = {}
        for addr in DEFAULT_INPUTS.keys():
            sheet_name, a1 = _split_sheet_address(str(addr))
            if sheet_name not in wb.sheetnames:
                raise KeyError(f"Workbook is missing sheet {sheet_name!r} for address {addr}")
            ws = ws_cache.get(sheet_name)
            if ws is None:
                ws = wb[sheet_name]
                ws_cache[sheet_name] = ws
            value = ws[a1].value
            updates[str(addr)] = 0 if value is None else value
        return updates
    finally:
        wb.close()


def _apply_range(
    ctx: EvalContext,
    *,
    shape: tuple[int, int],
    addresses: Sequence[str],
    values: object,
) -> RangeAssignment:
    rows, cols = shape
    updates: dict[str, CellValue] = {}
    flat: list[CellValue] = []
    if rows == 1 and cols == 1:
        flat = [values]  # scalar
    elif rows == 1 or cols == 1:
        if not isinstance(values, Sequence):
            raise TypeError('Expected a sequence for 1D range')
        flat = list(values)
        if len(flat) != len(addresses):
            raise ValueError(f'Expected {len(addresses)} values, got {len(flat)}')
    else:
        if not isinstance(values, Sequence):
            raise TypeError('Expected a sequence of sequences for 2D range')
        rows_values = list(values)
        if len(rows_values) != rows:
            raise ValueError(f'Expected {rows} rows, got {len(rows_values)}')
        for rv in rows_values:
            if not isinstance(rv, Sequence):
                raise TypeError('Expected a sequence of sequences for 2D range')
            row_list = list(rv)
            if len(row_list) != cols:
                raise ValueError(f'Expected {cols} columns, got {len(row_list)}')
            flat.extend(row_list)
    if len(flat) != len(addresses):
        raise ValueError(f'Expected {len(addresses)} values, got {len(flat)}')
    for addr, value in zip(addresses, flat):
        v = 0 if value is None else value
        updates[str(addr)] = v
    if updates:
        ctx.set_inputs(updates)
    return RangeAssignment(shape=shape, addresses=tuple(addresses))


def _apply_year_row_mapping(
    ctx: EvalContext,
    *,
    years: tuple[int, ...],
    year_to_addresses: dict[int, tuple[str, ...]],
    values_by_year: Mapping[int, CellValue],
    strict: bool = True,
) -> YearRowAssignment:
    applied: dict[int, tuple[str, ...]] = {}
    ignored: dict[int, CellValue] = {}
    updates: dict[str, CellValue] = {}
    for year, value in values_by_year.items():
        addrs = year_to_addresses.get(int(year))
        if addrs is None:
            if strict:
                raise KeyError(f"Year {year} is not in this table: {years}")
            ignored[int(year)] = value
            continue
        v = 0 if value is None else value
        for addr in addrs:
            updates[str(addr)] = v
        applied[int(year)] = tuple(addrs)
    if updates:
        ctx.set_inputs(updates)
    return YearRowAssignment(years=years, applied=applied, ignored=ignored)


def _apply_year_row_array(
    ctx: EvalContext,
    *,
    years: tuple[int, ...],
    year_to_addresses: dict[int, tuple[str, ...]],
    values: Sequence[CellValue],
    start_year: int,
    strict: bool = True,
) -> YearRowAssignment:
    if start_year not in year_to_addresses:
        raise KeyError(f"start_year {start_year} is not in this table: {years}")
    years_list = list(years)
    start_idx = years_list.index(start_year)
    remaining_years = years_list[start_idx:]
    if len(values) > len(remaining_years):
        raise ValueError(
            f"Too many values ({len(values)}) for table from {start_year}; "
            f"only {len(remaining_years)} years available"
        )
    expected = list(range(start_year, start_year + len(remaining_years)))
    if remaining_years != expected:
        raise ValueError(
            "Non-contiguous years; array mapping is disallowed for this table. "
            "Use dict-based mapping instead."
        )
    values_by_year = {start_year + i: values[i] for i in range(len(values))}
    return _apply_year_row_mapping(
        ctx, years=years, year_to_addresses=year_to_addresses, values_by_year=values_by_year, strict=strict
    )


def _apply_year_series_mapping(
    ctx: EvalContext,
    *,
    years: tuple[int, ...],
    year_to_address: dict[int, str],
    values_by_year: Mapping[int, CellValue],
    strict: bool = True,
) -> YearSeriesAssignment:
    applied: dict[int, str] = {}
    ignored: dict[int, CellValue] = {}
    updates: dict[str, CellValue] = {}
    for year, value in values_by_year.items():
        addr = year_to_address.get(int(year))
        if addr is None:
            if strict:
                raise KeyError(f"Year {year} is not in this series: {years}")
            ignored[int(year)] = value
            continue
        v = 0 if value is None else value
        updates[addr] = v
        applied[int(year)] = addr
    if updates:
        ctx.set_inputs(updates)
    return YearSeriesAssignment(years=years, applied=applied, ignored=ignored)


def _apply_year_series_array(
    ctx: EvalContext,
    *,
    years: tuple[int, ...],
    year_to_address: dict[int, str],
    values: Sequence[CellValue],
    start_year: int,
    strict: bool = True,
) -> YearSeriesAssignment:
    if start_year not in year_to_address:
        raise KeyError(f"start_year {start_year} is not in this series: {years}")
    years_list = list(years)
    start_idx = years_list.index(start_year)
    remaining_years = years_list[start_idx:]
    if len(values) > len(remaining_years):
        raise ValueError(
            f"Too many values ({len(values)}) for series from {start_year}; "
            f"only {len(remaining_years)} years available"
        )
    expected = list(range(start_year, start_year + len(remaining_years)))
    if remaining_years != expected:
        raise ValueError(
            "Non-contiguous years; array mapping is disallowed for this series. "
            "Use dict-based mapping instead."
        )
    values_by_year = {start_year + i: values[i] for i in range(len(values))}
    return _apply_year_series_mapping(
        ctx, years=years, year_to_address=year_to_address, values_by_year=values_by_year, strict=strict
    )


class LicDsfContext(EvalContext):
    __slots__ = ()

    def load_inputs_from_workbook(self, workbook_path: str) -> dict[str, CellValue]:
        updates = _read_inputs_from_workbook(workbook_path)
        if updates:
            self.set_inputs(updates)
        return updates

    def set_ext_debt_data_interest(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'Ext_Debt_Data!F384'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'Ext_Debt_Data!F384'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_ext_debt_data_nominal_value_pv_of_st_debt_locally_issued_debt(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023,), year_to_address={2023: 'Ext_Debt_Data!E382'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023,), year_to_address={2023: 'Ext_Debt_Data!E382'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_ext_debt_data_principal(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'Ext_Debt_Data!F383'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'Ext_Debt_Data!F383'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_1_basics_first_year_of_projections(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: "'Input 1 - Basics'!C18"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: "'Input 1 - Basics'!C18"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_current_account(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2025: "'Input 3 - Macro-Debt data(DMX)'!Y34", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z34", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA34", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB34", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC34", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD34", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE34", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF34", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG34", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH34", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI34", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ34", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK34", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL34", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM34", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN34", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO34", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP34", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ34", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR34"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2025: "'Input 3 - Macro-Debt data(DMX)'!Y34", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z34", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA34", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB34", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC34", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD34", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE34", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF34", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG34", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH34", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI34", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ34", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK34", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL34", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM34", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN34", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO34", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP34", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ34", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR34"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_debt_relief_non_multilateral_hipc(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X29", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y29", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z29", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA29", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB29", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC29", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD29", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE29", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF29", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG29", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH29", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI29", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ29", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK29", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL29", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM29", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN29", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO29", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP29", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ29", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR29"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X29", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y29", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z29", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA29", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB29", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC29", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD29", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE29", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF29", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG29", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH29", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI29", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ29", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK29", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL29", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM29", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN29", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO29", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP29", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ29", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR29"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_exports_of_goods_and_services(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2013: "'Input 3 - Macro-Debt data(DMX)'!M35", 2014: "'Input 3 - Macro-Debt data(DMX)'!N35", 2015: "'Input 3 - Macro-Debt data(DMX)'!O35", 2016: "'Input 3 - Macro-Debt data(DMX)'!P35", 2017: "'Input 3 - Macro-Debt data(DMX)'!Q35", 2018: "'Input 3 - Macro-Debt data(DMX)'!R35", 2019: "'Input 3 - Macro-Debt data(DMX)'!S35", 2020: "'Input 3 - Macro-Debt data(DMX)'!T35", 2021: "'Input 3 - Macro-Debt data(DMX)'!U35", 2022: "'Input 3 - Macro-Debt data(DMX)'!V35", 2023: "'Input 3 - Macro-Debt data(DMX)'!W35", 2024: "'Input 3 - Macro-Debt data(DMX)'!X35", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y35", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z35", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA35", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB35", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC35", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD35", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE35", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF35", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG35", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH35", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI35", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ35", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK35", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL35", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM35", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN35", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO35", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP35", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ35", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR35"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2013: "'Input 3 - Macro-Debt data(DMX)'!M35", 2014: "'Input 3 - Macro-Debt data(DMX)'!N35", 2015: "'Input 3 - Macro-Debt data(DMX)'!O35", 2016: "'Input 3 - Macro-Debt data(DMX)'!P35", 2017: "'Input 3 - Macro-Debt data(DMX)'!Q35", 2018: "'Input 3 - Macro-Debt data(DMX)'!R35", 2019: "'Input 3 - Macro-Debt data(DMX)'!S35", 2020: "'Input 3 - Macro-Debt data(DMX)'!T35", 2021: "'Input 3 - Macro-Debt data(DMX)'!U35", 2022: "'Input 3 - Macro-Debt data(DMX)'!V35", 2023: "'Input 3 - Macro-Debt data(DMX)'!W35", 2024: "'Input 3 - Macro-Debt data(DMX)'!X35", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y35", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z35", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA35", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB35", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC35", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD35", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE35", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF35", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG35", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH35", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI35", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ35", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK35", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL35", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM35", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN35", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO35", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP35", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ35", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR35"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_government_primary_expenditures_this_used_to_be_total_expenditure(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X24", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y24", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z24", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA24", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB24", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC24", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD24", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE24", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF24", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG24", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH24", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI24", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ24", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK24", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL24", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM24", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN24", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO24", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP24", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ24", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR24"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X24", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y24", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z24", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA24", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB24", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC24", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD24", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE24", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF24", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG24", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH24", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI24", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ24", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK24", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL24", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM24", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN24", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO24", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP24", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ24", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR24"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_government_grants(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W23", 2024: "'Input 3 - Macro-Debt data(DMX)'!X23", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y23", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z23", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA23", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB23", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC23", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD23", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE23", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF23", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG23", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH23", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI23", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ23", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK23", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL23", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM23", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN23", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO23", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP23", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ23", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR23"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W23", 2024: "'Input 3 - Macro-Debt data(DMX)'!X23", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y23", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z23", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA23", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB23", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC23", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD23", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE23", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF23", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG23", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH23", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI23", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ23", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK23", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL23", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM23", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN23", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO23", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP23", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ23", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR23"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_government_revenue_and_grants(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W22", 2024: "'Input 3 - Macro-Debt data(DMX)'!X22", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y22", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z22", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA22", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB22", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC22", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD22", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE22", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF22", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG22", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH22", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI22", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ22", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK22", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL22", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM22", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN22", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO22", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP22", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ22", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR22"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W22", 2024: "'Input 3 - Macro-Debt data(DMX)'!X22", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y22", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z22", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA22", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB22", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC22", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD22", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE22", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF22", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG22", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH22", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI22", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ22", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK22", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL22", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM22", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN22", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO22", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP22", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ22", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR22"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_gross_domestic_product_us_dollars(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2014: "'Input 3 - Macro-Debt data(DMX)'!N12", 2015: "'Input 3 - Macro-Debt data(DMX)'!O12", 2016: "'Input 3 - Macro-Debt data(DMX)'!P12", 2017: "'Input 3 - Macro-Debt data(DMX)'!Q12", 2018: "'Input 3 - Macro-Debt data(DMX)'!R12", 2019: "'Input 3 - Macro-Debt data(DMX)'!S12", 2020: "'Input 3 - Macro-Debt data(DMX)'!T12", 2021: "'Input 3 - Macro-Debt data(DMX)'!U12", 2022: "'Input 3 - Macro-Debt data(DMX)'!V12", 2023: "'Input 3 - Macro-Debt data(DMX)'!W12", 2024: "'Input 3 - Macro-Debt data(DMX)'!X12", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y12", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z12", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA12", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB12", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC12", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD12", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE12", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF12", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG12", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH12", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI12", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ12", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK12", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL12", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM12", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN12", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO12", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP12", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ12", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR12"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2014: "'Input 3 - Macro-Debt data(DMX)'!N12", 2015: "'Input 3 - Macro-Debt data(DMX)'!O12", 2016: "'Input 3 - Macro-Debt data(DMX)'!P12", 2017: "'Input 3 - Macro-Debt data(DMX)'!Q12", 2018: "'Input 3 - Macro-Debt data(DMX)'!R12", 2019: "'Input 3 - Macro-Debt data(DMX)'!S12", 2020: "'Input 3 - Macro-Debt data(DMX)'!T12", 2021: "'Input 3 - Macro-Debt data(DMX)'!U12", 2022: "'Input 3 - Macro-Debt data(DMX)'!V12", 2023: "'Input 3 - Macro-Debt data(DMX)'!W12", 2024: "'Input 3 - Macro-Debt data(DMX)'!X12", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y12", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z12", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA12", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB12", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC12", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD12", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE12", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF12", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG12", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH12", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI12", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ12", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK12", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL12", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM12", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN12", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO12", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP12", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ12", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR12"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ida_50y_loans(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X102", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y102", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z102", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA102", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB102", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC102", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD102", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE102", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF102", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG102", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH102", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI102", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ102", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK102", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL102", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM102", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN102", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO102", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP102", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ102", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR102"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X102", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y102", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z102", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA102", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB102", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC102", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD102", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE102", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF102", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG102", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH102", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI102", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ102", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK102", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL102", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM102", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN102", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO102", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP102", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ102", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR102"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ida_sml(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X103", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y103", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z103", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA103", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB103", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC103", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD103", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE103", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF103", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG103", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH103", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI103", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ103", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK103", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL103", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM103", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN103", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO103", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP103", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ103", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR103"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X103", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y103", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z103", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA103", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB103", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC103", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD103", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE103", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF103", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG103", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH103", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI103", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ103", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK103", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL103", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM103", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN103", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO103", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP103", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ103", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR103"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ida_new_40_year_credits(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X104", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y104"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X104", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y104"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ida_new_60_year_credits(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X107"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X107"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ida_new_blend(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X106"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X106"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ida_new_regular(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X105", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y105"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X105", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y105"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_imports_of_goods_and_services_enter_as_a_positive_number(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2025: "'Input 3 - Macro-Debt data(DMX)'!Y38", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z38", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA38", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB38", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC38", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD38", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE38", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF38", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG38", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH38", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI38", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ38", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK38", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL38", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM38", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN38", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO38", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP38", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ38", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR38"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2025: "'Input 3 - Macro-Debt data(DMX)'!Y38", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z38", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA38", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB38", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC38", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD38", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE38", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF38", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG38", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH38", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI38", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ38", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK38", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL38", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM38", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN38", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO38", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP38", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ38", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR38"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_multilateral1(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044, 2045, 2046, 2047, 2048, 2049, 2050, 2051, 2052, 2053, 2054, 2055, 2056, 2057, 2058, 2059, 2060, 2061, 2062, 2063, 2064, 2065, 2066, 2067, 2068), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W68", 2024: "'Input 3 - Macro-Debt data(DMX)'!X68", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y68", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z68", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA68", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB68", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC68", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD68", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE68", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF68", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG68", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH68", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI68", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ68", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK68", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL68", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM68", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN68", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO68", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP68", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ68", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR68", 2045: "'Input 3 - Macro-Debt data(DMX)'!AS68", 2046: "'Input 3 - Macro-Debt data(DMX)'!AT68", 2047: "'Input 3 - Macro-Debt data(DMX)'!AU68", 2048: "'Input 3 - Macro-Debt data(DMX)'!AV68", 2049: "'Input 3 - Macro-Debt data(DMX)'!AW68", 2050: "'Input 3 - Macro-Debt data(DMX)'!AX68", 2051: "'Input 3 - Macro-Debt data(DMX)'!AY68", 2052: "'Input 3 - Macro-Debt data(DMX)'!AZ68", 2053: "'Input 3 - Macro-Debt data(DMX)'!BA68", 2054: "'Input 3 - Macro-Debt data(DMX)'!BB68", 2055: "'Input 3 - Macro-Debt data(DMX)'!BC68", 2056: "'Input 3 - Macro-Debt data(DMX)'!BD68", 2057: "'Input 3 - Macro-Debt data(DMX)'!BE68", 2058: "'Input 3 - Macro-Debt data(DMX)'!BF68", 2059: "'Input 3 - Macro-Debt data(DMX)'!BG68", 2060: "'Input 3 - Macro-Debt data(DMX)'!BH68", 2061: "'Input 3 - Macro-Debt data(DMX)'!BI68", 2062: "'Input 3 - Macro-Debt data(DMX)'!BJ68", 2063: "'Input 3 - Macro-Debt data(DMX)'!BK68", 2064: "'Input 3 - Macro-Debt data(DMX)'!BL68", 2065: "'Input 3 - Macro-Debt data(DMX)'!BM68", 2066: "'Input 3 - Macro-Debt data(DMX)'!BN68", 2067: "'Input 3 - Macro-Debt data(DMX)'!BO68", 2068: "'Input 3 - Macro-Debt data(DMX)'!BP68"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044, 2045, 2046, 2047, 2048, 2049, 2050, 2051, 2052, 2053, 2054, 2055, 2056, 2057, 2058, 2059, 2060, 2061, 2062, 2063, 2064, 2065, 2066, 2067, 2068), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W68", 2024: "'Input 3 - Macro-Debt data(DMX)'!X68", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y68", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z68", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA68", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB68", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC68", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD68", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE68", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF68", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG68", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH68", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI68", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ68", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK68", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL68", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM68", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN68", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO68", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP68", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ68", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR68", 2045: "'Input 3 - Macro-Debt data(DMX)'!AS68", 2046: "'Input 3 - Macro-Debt data(DMX)'!AT68", 2047: "'Input 3 - Macro-Debt data(DMX)'!AU68", 2048: "'Input 3 - Macro-Debt data(DMX)'!AV68", 2049: "'Input 3 - Macro-Debt data(DMX)'!AW68", 2050: "'Input 3 - Macro-Debt data(DMX)'!AX68", 2051: "'Input 3 - Macro-Debt data(DMX)'!AY68", 2052: "'Input 3 - Macro-Debt data(DMX)'!AZ68", 2053: "'Input 3 - Macro-Debt data(DMX)'!BA68", 2054: "'Input 3 - Macro-Debt data(DMX)'!BB68", 2055: "'Input 3 - Macro-Debt data(DMX)'!BC68", 2056: "'Input 3 - Macro-Debt data(DMX)'!BD68", 2057: "'Input 3 - Macro-Debt data(DMX)'!BE68", 2058: "'Input 3 - Macro-Debt data(DMX)'!BF68", 2059: "'Input 3 - Macro-Debt data(DMX)'!BG68", 2060: "'Input 3 - Macro-Debt data(DMX)'!BH68", 2061: "'Input 3 - Macro-Debt data(DMX)'!BI68", 2062: "'Input 3 - Macro-Debt data(DMX)'!BJ68", 2063: "'Input 3 - Macro-Debt data(DMX)'!BK68", 2064: "'Input 3 - Macro-Debt data(DMX)'!BL68", 2065: "'Input 3 - Macro-Debt data(DMX)'!BM68", 2066: "'Input 3 - Macro-Debt data(DMX)'!BN68", 2067: "'Input 3 - Macro-Debt data(DMX)'!BO68", 2068: "'Input 3 - Macro-Debt data(DMX)'!BP68"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_national_currency_per_u_s_dollar_e_o_p(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W19", 2024: "'Input 3 - Macro-Debt data(DMX)'!X19", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y19", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z19", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA19", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB19", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC19", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD19", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE19", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF19", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG19", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH19", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI19", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ19", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK19", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL19", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM19", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN19", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO19", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP19", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ19", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR19"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W19", 2024: "'Input 3 - Macro-Debt data(DMX)'!X19", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y19", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z19", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA19", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB19", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC19", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD19", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE19", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF19", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG19", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH19", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI19", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ19", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK19", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL19", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM19", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN19", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO19", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP19", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ19", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR19"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_national_currency_per_u_s_dollar_p_a(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W20", 2024: "'Input 3 - Macro-Debt data(DMX)'!X20", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y20", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z20", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA20", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB20", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC20", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD20", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE20", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF20", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG20", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH20", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI20", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ20", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK20", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL20", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM20", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN20", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO20", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP20", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ20", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR20"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W20", 2024: "'Input 3 - Macro-Debt data(DMX)'!X20", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y20", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z20", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA20", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB20", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC20", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD20", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE20", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF20", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG20", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH20", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI20", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ20", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK20", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL20", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM20", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN20", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO20", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP20", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ20", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR20"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_new_gross_disbursement_central_bank(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X147", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y147", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z147", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA147", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB147", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC147", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD147", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE147", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF147", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG147", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH147", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI147", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ147", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK147", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL147", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM147", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN147", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO147", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP147", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ147", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR147"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X147", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y147", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z147", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA147", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB147", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC147", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD147", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE147", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF147", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG147", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH147", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI147", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ147", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK147", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL147", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM147", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN147", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO147", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP147", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ147", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR147"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_other_debt_creating_or_reducing_flow_please_specify(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X30", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y30", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z30", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA30", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB30", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC30", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD30", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE30", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF30", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG30", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH30", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI30", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ30", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK30", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL30", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM30", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN30", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO30", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP30", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ30", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR30"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X30", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y30", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z30", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA30", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB30", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC30", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD30", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE30", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF30", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG30", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH30", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI30", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ30", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK30", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL30", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM30", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN30", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO30", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP30", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ30", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR30"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_outstanding_of_existing_debt_in_local_currency(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W161"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W161"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ppg_mlt_external_debt_outstanding(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W51"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W51"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ppg_st_external_debt_outstanding(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W52", 2024: "'Input 3 - Macro-Debt data(DMX)'!X52", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y52", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z52", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA52", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB52", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC52", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD52", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE52", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF52", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG52", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH52", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI52", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ52", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK52", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL52", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM52", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN52", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO52", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP52", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ52", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR52"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W52", 2024: "'Input 3 - Macro-Debt data(DMX)'!X52", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y52", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z52", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA52", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB52", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC52", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD52", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE52", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF52", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG52", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH52", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI52", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ52", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK52", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL52", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM52", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN52", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO52", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP52", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ52", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR52"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ppg_total_external_debt_amortization_due(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W54"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W54"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_ppg_external_debt_interest_due(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W53"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023,), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W53"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_private_mlt_external_debt_amortization_due(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W60", 2024: "'Input 3 - Macro-Debt data(DMX)'!X60"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W60", 2024: "'Input 3 - Macro-Debt data(DMX)'!X60"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_private_external_debt_interest_due(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W59", 2024: "'Input 3 - Macro-Debt data(DMX)'!X59", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y59", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z59", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA59", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB59", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC59", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD59", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE59", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF59", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG59", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH59", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI59", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ59", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK59", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL59", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM59", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN59", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO59", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP59", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ59", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR59"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W59", 2024: "'Input 3 - Macro-Debt data(DMX)'!X59", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y59", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z59", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA59", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB59", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC59", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD59", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE59", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF59", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG59", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH59", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI59", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ59", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK59", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL59", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM59", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN59", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO59", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP59", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ59", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR59"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_private_sector_mlt_external_debt_outstanding(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W57", 2024: "'Input 3 - Macro-Debt data(DMX)'!X57", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y57", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z57", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA57", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB57", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC57", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD57", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE57", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF57", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG57", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH57", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI57", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ57", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK57", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL57", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM57", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN57", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO57", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP57", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ57", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR57"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2023: "'Input 3 - Macro-Debt data(DMX)'!W57", 2024: "'Input 3 - Macro-Debt data(DMX)'!X57", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y57", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z57", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA57", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB57", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC57", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD57", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE57", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF57", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG57", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH57", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI57", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ57", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK57", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL57", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM57", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN57", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO57", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP57", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ57", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR57"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_private_sector_st_external_debt_outstanding(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2022: "'Input 3 - Macro-Debt data(DMX)'!V58", 2023: "'Input 3 - Macro-Debt data(DMX)'!W58", 2024: "'Input 3 - Macro-Debt data(DMX)'!X58", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y58", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z58", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA58", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB58", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC58", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD58", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE58", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF58", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG58", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH58", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI58", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ58", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK58", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL58", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM58", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN58", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO58", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP58", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ58", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR58"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2022: "'Input 3 - Macro-Debt data(DMX)'!V58", 2023: "'Input 3 - Macro-Debt data(DMX)'!W58", 2024: "'Input 3 - Macro-Debt data(DMX)'!X58", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y58", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z58", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA58", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB58", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC58", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD58", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE58", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF58", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG58", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH58", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI58", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ58", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK58", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL58", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM58", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN58", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO58", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP58", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ58", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR58"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_privatization_proceeds(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X27", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y27", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z27", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA27", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB27", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC27", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD27", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE27", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF27", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG27", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH27", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI27", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ27", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK27", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL27", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM27", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN27", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO27", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP27", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ27", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR27"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X27", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y27", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z27", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA27", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB27", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC27", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD27", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE27", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF27", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG27", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH27", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI27", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ27", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK27", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL27", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM27", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN27", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO27", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP27", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ27", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR27"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_real_gross_domestic_product(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2013: "'Input 3 - Macro-Debt data(DMX)'!M13", 2014: "'Input 3 - Macro-Debt data(DMX)'!N13", 2015: "'Input 3 - Macro-Debt data(DMX)'!O13", 2016: "'Input 3 - Macro-Debt data(DMX)'!P13", 2017: "'Input 3 - Macro-Debt data(DMX)'!Q13", 2018: "'Input 3 - Macro-Debt data(DMX)'!R13", 2019: "'Input 3 - Macro-Debt data(DMX)'!S13", 2020: "'Input 3 - Macro-Debt data(DMX)'!T13", 2021: "'Input 3 - Macro-Debt data(DMX)'!U13", 2022: "'Input 3 - Macro-Debt data(DMX)'!V13", 2023: "'Input 3 - Macro-Debt data(DMX)'!W13", 2024: "'Input 3 - Macro-Debt data(DMX)'!X13", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y13", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z13", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA13", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB13", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC13", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD13", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE13", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF13", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG13", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH13", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI13", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ13", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK13", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL13", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM13", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN13", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO13", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP13", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ13", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR13"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2013: "'Input 3 - Macro-Debt data(DMX)'!M13", 2014: "'Input 3 - Macro-Debt data(DMX)'!N13", 2015: "'Input 3 - Macro-Debt data(DMX)'!O13", 2016: "'Input 3 - Macro-Debt data(DMX)'!P13", 2017: "'Input 3 - Macro-Debt data(DMX)'!Q13", 2018: "'Input 3 - Macro-Debt data(DMX)'!R13", 2019: "'Input 3 - Macro-Debt data(DMX)'!S13", 2020: "'Input 3 - Macro-Debt data(DMX)'!T13", 2021: "'Input 3 - Macro-Debt data(DMX)'!U13", 2022: "'Input 3 - Macro-Debt data(DMX)'!V13", 2023: "'Input 3 - Macro-Debt data(DMX)'!W13", 2024: "'Input 3 - Macro-Debt data(DMX)'!X13", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y13", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z13", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA13", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB13", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC13", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD13", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE13", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF13", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG13", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH13", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI13", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ13", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK13", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL13", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM13", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN13", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO13", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP13", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ13", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR13"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_recognition_of_contingent_liabilities_e_g_bank_recapitalization(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X28", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y28", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z28", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA28", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB28", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC28", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD28", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE28", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF28", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG28", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH28", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI28", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ28", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK28", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL28", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM28", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN28", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO28", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP28", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ28", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR28"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X28", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y28", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z28", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA28", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB28", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC28", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD28", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE28", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF28", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG28", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH28", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI28", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ28", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK28", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL28", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM28", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN28", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO28", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP28", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ28", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR28"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_3_macro_debt_data_dmx_total_principal_payment(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X95", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y95", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z95", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA95", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB95", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC95", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD95", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE95", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF95", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG95", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH95", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI95", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ95", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK95", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL95", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM95", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN95", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO95", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP95", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ95", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR95"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 3 - Macro-Debt data(DMX)'!X95", 2025: "'Input 3 - Macro-Debt data(DMX)'!Y95", 2026: "'Input 3 - Macro-Debt data(DMX)'!Z95", 2027: "'Input 3 - Macro-Debt data(DMX)'!AA95", 2028: "'Input 3 - Macro-Debt data(DMX)'!AB95", 2029: "'Input 3 - Macro-Debt data(DMX)'!AC95", 2030: "'Input 3 - Macro-Debt data(DMX)'!AD95", 2031: "'Input 3 - Macro-Debt data(DMX)'!AE95", 2032: "'Input 3 - Macro-Debt data(DMX)'!AF95", 2033: "'Input 3 - Macro-Debt data(DMX)'!AG95", 2034: "'Input 3 - Macro-Debt data(DMX)'!AH95", 2035: "'Input 3 - Macro-Debt data(DMX)'!AI95", 2036: "'Input 3 - Macro-Debt data(DMX)'!AJ95", 2037: "'Input 3 - Macro-Debt data(DMX)'!AK95", 2038: "'Input 3 - Macro-Debt data(DMX)'!AL95", 2039: "'Input 3 - Macro-Debt data(DMX)'!AM95", 2040: "'Input 3 - Macro-Debt data(DMX)'!AN95", 2041: "'Input 3 - Macro-Debt data(DMX)'!AO95", 2042: "'Input 3 - Macro-Debt data(DMX)'!AP95", 2043: "'Input 3 - Macro-Debt data(DMX)'!AQ95", 2044: "'Input 3 - Macro-Debt data(DMX)'!AR95"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_4_external_financing_ida_50y_loans(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2031, 2032, 2033), year_to_address={2031: "'Input 4 - External Financing'!S71", 2032: "'Input 4 - External Financing'!T71", 2033: "'Input 4 - External Financing'!U71"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2031, 2032, 2033), year_to_address={2031: "'Input 4 - External Financing'!S71", 2032: "'Input 4 - External Financing'!T71", 2033: "'Input 4 - External Financing'!U71"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_4_external_financing_ida_sml(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2027,), year_to_address={2027: "'Input 4 - External Financing'!O70"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2027,), year_to_address={2027: "'Input 4 - External Financing'!O70"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_4_external_financing_ida_blend(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2026,), year_to_address={2026: "'Input 4 - External Financing'!N69"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2026,), year_to_address={2026: "'Input 4 - External Financing'!N69"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_4_external_financing_ida_small_economy(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2031: "'Input 4 - External Financing'!S67", 2032: "'Input 4 - External Financing'!T67", 2033: "'Input 4 - External Financing'!U67", 2034: "'Input 4 - External Financing'!V67", 2035: "'Input 4 - External Financing'!W67", 2036: "'Input 4 - External Financing'!X67", 2037: "'Input 4 - External Financing'!Y67", 2038: "'Input 4 - External Financing'!Z67", 2039: "'Input 4 - External Financing'!AA67", 2040: "'Input 4 - External Financing'!AB67", 2041: "'Input 4 - External Financing'!AC67", 2042: "'Input 4 - External Financing'!AD67", 2043: "'Input 4 - External Financing'!AE67", 2044: "'Input 4 - External Financing'!AF67"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2031: "'Input 4 - External Financing'!S67", 2032: "'Input 4 - External Financing'!T67", 2033: "'Input 4 - External Financing'!U67", 2034: "'Input 4 - External Financing'!V67", 2035: "'Input 4 - External Financing'!W67", 2036: "'Input 4 - External Financing'!X67", 2037: "'Input 4 - External Financing'!Y67", 2038: "'Input 4 - External Financing'!Z67", 2039: "'Input 4 - External Financing'!AA67", 2040: "'Input 4 - External Financing'!AB67", 2041: "'Input 4 - External Financing'!AC67", 2042: "'Input 4 - External Financing'!AD67", 2043: "'Input 4 - External Financing'!AE67", 2044: "'Input 4 - External Financing'!AF67"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_4_external_financing_ida_new_40_year_credits(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2026,), year_to_address={2026: "'Input 4 - External Financing'!N14"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2026,), year_to_address={2026: "'Input 4 - External Financing'!N14"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_bonds_1_to_3_years_fx(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I20", 2025: "'Input 5 - Local-debt Financing'!J20", 2026: "'Input 5 - Local-debt Financing'!K20", 2027: "'Input 5 - Local-debt Financing'!L20", 2028: "'Input 5 - Local-debt Financing'!M20", 2029: "'Input 5 - Local-debt Financing'!N20"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I20", 2025: "'Input 5 - Local-debt Financing'!J20", 2026: "'Input 5 - Local-debt Financing'!K20", 2027: "'Input 5 - Local-debt Financing'!L20", 2028: "'Input 5 - Local-debt Financing'!M20", 2029: "'Input 5 - Local-debt Financing'!N20"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_bonds_1_to_3_years_lc(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I16", 2025: "'Input 5 - Local-debt Financing'!J16", 2026: "'Input 5 - Local-debt Financing'!K16", 2027: "'Input 5 - Local-debt Financing'!L16", 2028: "'Input 5 - Local-debt Financing'!M16", 2029: "'Input 5 - Local-debt Financing'!N16"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I16", 2025: "'Input 5 - Local-debt Financing'!J16", 2026: "'Input 5 - Local-debt Financing'!K16", 2027: "'Input 5 - Local-debt Financing'!L16", 2028: "'Input 5 - Local-debt Financing'!M16", 2029: "'Input 5 - Local-debt Financing'!N16"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_bonds_4_to_7_years_fx(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I21", 2025: "'Input 5 - Local-debt Financing'!J21", 2026: "'Input 5 - Local-debt Financing'!K21", 2027: "'Input 5 - Local-debt Financing'!L21", 2028: "'Input 5 - Local-debt Financing'!M21", 2029: "'Input 5 - Local-debt Financing'!N21"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I21", 2025: "'Input 5 - Local-debt Financing'!J21", 2026: "'Input 5 - Local-debt Financing'!K21", 2027: "'Input 5 - Local-debt Financing'!L21", 2028: "'Input 5 - Local-debt Financing'!M21", 2029: "'Input 5 - Local-debt Financing'!N21"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_bonds_4_to_7_years_lc(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I17", 2025: "'Input 5 - Local-debt Financing'!J17", 2026: "'Input 5 - Local-debt Financing'!K17", 2027: "'Input 5 - Local-debt Financing'!L17", 2028: "'Input 5 - Local-debt Financing'!M17", 2029: "'Input 5 - Local-debt Financing'!N17"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I17", 2025: "'Input 5 - Local-debt Financing'!J17", 2026: "'Input 5 - Local-debt Financing'!K17", 2027: "'Input 5 - Local-debt Financing'!L17", 2028: "'Input 5 - Local-debt Financing'!M17", 2029: "'Input 5 - Local-debt Financing'!N17"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_bonds_beyond_7_years_fx(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I22", 2025: "'Input 5 - Local-debt Financing'!J22", 2026: "'Input 5 - Local-debt Financing'!K22", 2027: "'Input 5 - Local-debt Financing'!L22", 2028: "'Input 5 - Local-debt Financing'!M22", 2029: "'Input 5 - Local-debt Financing'!N22"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I22", 2025: "'Input 5 - Local-debt Financing'!J22", 2026: "'Input 5 - Local-debt Financing'!K22", 2027: "'Input 5 - Local-debt Financing'!L22", 2028: "'Input 5 - Local-debt Financing'!M22", 2029: "'Input 5 - Local-debt Financing'!N22"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_bonds_beyond_7_years_lc(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I18", 2025: "'Input 5 - Local-debt Financing'!J18", 2026: "'Input 5 - Local-debt Financing'!K18", 2027: "'Input 5 - Local-debt Financing'!L18", 2028: "'Input 5 - Local-debt Financing'!M18", 2029: "'Input 5 - Local-debt Financing'!N18"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I18", 2025: "'Input 5 - Local-debt Financing'!J18", 2026: "'Input 5 - Local-debt Financing'!K18", 2027: "'Input 5 - Local-debt Financing'!L18", 2028: "'Input 5 - Local-debt Financing'!M18", 2029: "'Input 5 - Local-debt Financing'!N18"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_central_bank_financing(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I10", 2025: "'Input 5 - Local-debt Financing'!J10", 2026: "'Input 5 - Local-debt Financing'!K10", 2027: "'Input 5 - Local-debt Financing'!L10", 2028: "'Input 5 - Local-debt Financing'!M10", 2029: "'Input 5 - Local-debt Financing'!N10"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I10", 2025: "'Input 5 - Local-debt Financing'!J10", 2026: "'Input 5 - Local-debt Financing'!K10", 2027: "'Input 5 - Local-debt Financing'!L10", 2028: "'Input 5 - Local-debt Financing'!M10", 2029: "'Input 5 - Local-debt Financing'!N10"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_t_bills_denominated_in_foreign_currency(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I13", 2025: "'Input 5 - Local-debt Financing'!J13", 2026: "'Input 5 - Local-debt Financing'!K13", 2027: "'Input 5 - Local-debt Financing'!L13", 2028: "'Input 5 - Local-debt Financing'!M13", 2029: "'Input 5 - Local-debt Financing'!N13"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I13", 2025: "'Input 5 - Local-debt Financing'!J13", 2026: "'Input 5 - Local-debt Financing'!K13", 2027: "'Input 5 - Local-debt Financing'!L13", 2028: "'Input 5 - Local-debt Financing'!M13", 2029: "'Input 5 - Local-debt Financing'!N13"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_5_local_debt_financing_t_bills_denominated_in_local_currency(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I12", 2025: "'Input 5 - Local-debt Financing'!J12", 2026: "'Input 5 - Local-debt Financing'!K12", 2027: "'Input 5 - Local-debt Financing'!L12", 2028: "'Input 5 - Local-debt Financing'!M12", 2029: "'Input 5 - Local-debt Financing'!N12"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029), year_to_address={2024: "'Input 5 - Local-debt Financing'!I12", 2025: "'Input 5 - Local-debt Financing'!J12", 2026: "'Input 5 - Local-debt Financing'!K12", 2027: "'Input 5 - Local-debt Financing'!L12", 2028: "'Input 5 - Local-debt Financing'!M12", 2029: "'Input 5 - Local-debt Financing'!N12"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_input_8_sdr_sdr_interest_rate(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 8 - SDR'!C14", 2025: "'Input 8 - SDR'!D14", 2026: "'Input 8 - SDR'!E14", 2027: "'Input 8 - SDR'!F14", 2028: "'Input 8 - SDR'!G14", 2029: "'Input 8 - SDR'!H14", 2030: "'Input 8 - SDR'!I14", 2031: "'Input 8 - SDR'!J14", 2032: "'Input 8 - SDR'!K14", 2033: "'Input 8 - SDR'!L14", 2034: "'Input 8 - SDR'!M14", 2035: "'Input 8 - SDR'!N14", 2036: "'Input 8 - SDR'!O14", 2037: "'Input 8 - SDR'!P14", 2038: "'Input 8 - SDR'!Q14", 2039: "'Input 8 - SDR'!R14", 2040: "'Input 8 - SDR'!S14", 2041: "'Input 8 - SDR'!T14", 2042: "'Input 8 - SDR'!U14", 2043: "'Input 8 - SDR'!V14", 2044: "'Input 8 - SDR'!W14"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044), year_to_address={2024: "'Input 8 - SDR'!C14", 2025: "'Input 8 - SDR'!D14", 2026: "'Input 8 - SDR'!E14", 2027: "'Input 8 - SDR'!F14", 2028: "'Input 8 - SDR'!G14", 2029: "'Input 8 - SDR'!H14", 2030: "'Input 8 - SDR'!I14", 2031: "'Input 8 - SDR'!J14", 2032: "'Input 8 - SDR'!K14", 2033: "'Input 8 - SDR'!L14", 2034: "'Input 8 - SDR'!M14", 2035: "'Input 8 - SDR'!N14", 2036: "'Input 8 - SDR'!O14", 2037: "'Input 8 - SDR'!P14", 2038: "'Input 8 - SDR'!Q14", 2039: "'Input 8 - SDR'!R14", 2040: "'Input 8 - SDR'!S14", 2041: "'Input 8 - SDR'!T14", 2042: "'Input 8 - SDR'!U14", 2043: "'Input 8 - SDR'!V14", 2044: "'Input 8 - SDR'!W14"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_stress_alternative_scenario_1_key_variables_at_historical_average(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: "'PV Stress'!D4"},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: "'PV Stress'!D4"},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_g00209(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D40'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D40'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D9'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D9'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_2(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D674'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D674'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_3(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D700'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D700'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_4(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D726'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D726'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_5(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D648'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D648'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_6(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D622'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D622'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_7(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D362'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D362'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_8(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D492'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D492'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_9(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D77'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D77'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_10(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D102'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D102'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_11(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D51'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D51'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_12(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D126'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D126'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_13(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D198'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D198'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_14(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D174'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D174'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_15(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D150'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D150'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_16(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D232'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D232'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_17(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D258'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D258'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_18(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D518'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D518'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_19(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D544'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D544'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_20(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D570'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D570'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_21(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D596'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D596'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_22(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D284'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D284'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_23(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D310'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D310'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_24(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D336'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D336'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_25(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D388'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D388'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_26(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D414'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D414'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_27(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D440'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D440'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_base_28(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D466'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D466'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_pv_base_ida_regular(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearSeriesAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_series_mapping(
                self, years=(2024,), year_to_address={2024: 'PV_Base!D49'},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-series inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_series_array(
            self, years=(2024,), year_to_address={2024: 'PV_Base!D49'},
            values=values, start_year=start_year, strict=strict,
        )

    def set_blend_floating_calculations_wb_g00002(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!D5"], values=values
        )

    def set_blend_floating_calculations_wb_g00003(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!M10"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_1_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K10"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_10_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K19"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_12_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K20"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_15_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K21"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_2_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K11"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_20_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K22"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_25_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K23"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_3_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K12"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_30_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K24"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_4_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K13"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_5_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K14"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_6_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K15"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_7_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K16"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_8_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K17"], values=values
        )

    def set_blend_floating_calculations_wb_sheet_9_year(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!K18"], values=values
        )

    def set_blend_floating_calculations_wb_ida_new_blend_floating(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'BLEND floating calculations WB'!C6"], values=values
        )

    def set_input_1_basics_discount_rate(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 1 - Basics'!C25"], values=values
        )

    def set_input_4_external_financing_com3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G40"], values=values
        )

    def set_input_4_external_financing_com3_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F40"], values=values
        )

    def set_input_4_external_financing_com3_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H40"], values=values
        )

    def set_input_4_external_financing_com4(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G41"], values=values
        )

    def set_input_4_external_financing_com4_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F41"], values=values
        )

    def set_input_4_external_financing_com4_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H41"], values=values
        )

    def set_input_4_external_financing_com5(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G42"], values=values
        )

    def set_input_4_external_financing_com5_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F42"], values=values
        )

    def set_input_4_external_financing_com5_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H42"], values=values
        )

    def set_input_4_external_financing_commecial_bank(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G39"], values=values
        )

    def set_input_4_external_financing_commecial_bank_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F39"], values=values
        )

    def set_input_4_external_financing_commecial_bank_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H39"], values=values
        )

    def set_input_4_external_financing_eurobond(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G38"], values=values
        )

    def set_input_4_external_financing_eurobond_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F38"], values=values
        )

    def set_input_4_external_financing_eurobond_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H38"], values=values
        )

    def set_input_4_external_financing_export_credit_agencies(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G26"], values=values
        )

    def set_input_4_external_financing_export_credit_agencies_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F26"], values=values
        )

    def set_input_4_external_financing_export_credit_agencies_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H26"], values=values
        )

    def set_input_4_external_financing_export_import_bank_of_npc(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G32"], values=values
        )

    def set_input_4_external_financing_export_import_bank_of_npc_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F32"], values=values
        )

    def set_input_4_external_financing_export_import_bank_of_npc_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H32"], values=values
        )

    def set_input_4_external_financing_ida_50y_loans_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!D71"], values=values
        )

    def set_input_4_external_financing_ida_50y_loans_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E71"], values=values
        )

    def set_input_4_external_financing_ida_50y_loans_4(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F71"], values=values
        )

    def set_input_4_external_financing_ida_sml_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!D70"], values=values
        )

    def set_input_4_external_financing_ida_sml_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E70"], values=values
        )

    def set_input_4_external_financing_ida_sml_4(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F70"], values=values
        )

    def set_input_4_external_financing_ida_blend_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E69"], values=values
        )

    def set_input_4_external_financing_ida_blend_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F69"], values=values
        )

    def set_input_4_external_financing_ida_regular(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!D68"], values=values
        )

    def set_input_4_external_financing_ida_regular_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E68"], values=values
        )

    def set_input_4_external_financing_ida_regular_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F68"], values=values
        )

    def set_input_4_external_financing_ida_small_economy_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E67"], values=values
        )

    def set_input_4_external_financing_ida_small_economy_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F67"], values=values
        )

    def set_input_4_external_financing_ida_new_40_year_credits_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!D72"], values=values
        )

    def set_input_4_external_financing_ida_new_40_year_credits_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E72"], values=values
        )

    def set_input_4_external_financing_ida_new_40_year_credits_4(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F72"], values=values
        )

    def set_input_4_external_financing_ida_new_60_year_credits(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E75"], values=values
        )

    def set_input_4_external_financing_ida_new_60_year_credits_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F75"], values=values
        )

    def set_input_4_external_financing_ida_new_blend_also_enter(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!AG74"], values=values
        )

    def set_input_4_external_financing_ida_new_blend_also_enter_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E74"], values=values
        )

    def set_input_4_external_financing_ida_new_blend_also_enter_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F74"], values=values
        )

    def set_input_4_external_financing_ida_new_regular(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!E73"], values=values
        )

    def set_input_4_external_financing_ida_new_regular_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F73"], values=values
        )

    def set_input_4_external_financing_imf(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G10"], values=values
        )

    def set_input_4_external_financing_imf_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F10"], values=values
        )

    def set_input_4_external_financing_imf_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H10"], values=values
        )

    def set_input_4_external_financing_multi1(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G18"], values=values
        )

    def set_input_4_external_financing_multi1_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F18"], values=values
        )

    def set_input_4_external_financing_multi1_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H18"], values=values
        )

    def set_input_4_external_financing_multi2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G19"], values=values
        )

    def set_input_4_external_financing_multi2_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F19"], values=values
        )

    def set_input_4_external_financing_multi2_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H19"], values=values
        )

    def set_input_4_external_financing_npc2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G33"], values=values
        )

    def set_input_4_external_financing_npc2_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F33"], values=values
        )

    def set_input_4_external_financing_npc2_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H33"], values=values
        )

    def set_input_4_external_financing_npc3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G34"], values=values
        )

    def set_input_4_external_financing_npc3_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F34"], values=values
        )

    def set_input_4_external_financing_npc3_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H34"], values=values
        )

    def set_input_4_external_financing_npc4(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G35"], values=values
        )

    def set_input_4_external_financing_npc4_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F35"], values=values
        )

    def set_input_4_external_financing_npc4_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H35"], values=values
        )

    def set_input_4_external_financing_npc5(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G36"], values=values
        )

    def set_input_4_external_financing_npc5_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F36"], values=values
        )

    def set_input_4_external_financing_npc5_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H36"], values=values
        )

    def set_input_4_external_financing_oth_multi1(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G21"], values=values
        )

    def set_input_4_external_financing_oth_multi1_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F21"], values=values
        )

    def set_input_4_external_financing_oth_multi1_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H21"], values=values
        )

    def set_input_4_external_financing_oth_multi2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G22"], values=values
        )

    def set_input_4_external_financing_oth_multi2_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F22"], values=values
        )

    def set_input_4_external_financing_oth_multi2_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H22"], values=values
        )

    def set_input_4_external_financing_oth_multi3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!G23"], values=values
        )

    def set_input_4_external_financing_oth_multi3_2(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F23"], values=values
        )

    def set_input_4_external_financing_oth_multi3_3(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!H23"], values=values
        )

    def set_input_4_external_financing_ppg_st_external_debt(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 4 - External Financing'!F45"], values=values
        )

    def set_input_5_local_debt_financing_g00191(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 5 - Local-debt Financing'!C78"], values=values
        )

    def set_input_6_optional_standard_test_current_transfers_to_gdp_and_fdi_to_gdp_ratios_set_to_their_historical_average_minus_one_sd_or_baseline_projection_minus_one_sd_whichever_is_lower_in_the_second_and_third_years_of_the_projection_period(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 6(optional)-Standard Test'!C29"], values=values
        )

    def set_input_6_optional_standard_test_nominal_export_growth_in_usd_set_to_its_historical_average_minus_one_sd_or_baseline_projection_minus_one_sd_whichever_is_lower_in_the_second_and_third_years_of_the_projection_period(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 6(optional)-Standard Test'!C25"], values=values
        )

    def set_input_6_optional_standard_test_other_flows_fdi_shock_of_standard_deviations(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 6(optional)-Standard Test'!C32"], values=values
        )

    def set_input_6_optional_standard_test_real_gdp_growth_set_to_its_historical_average_minus_one_sd_or_baseline_projection_minus_one_sd_whichever_is_lower_for_the_second_and_third_years_of_the_projection_period(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 6(optional)-Standard Test'!C17"], values=values
        )

    def set_input_8_sdr_sdr_allocation_in_million_of_usd(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 8 - SDR'!B6"], values=values
        )

    def set_input_8_sdr_sdr_holdings_in_million_of_usd(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=["'Input 8 - SDR'!B7"], values=values
        )

    def set_start_debt_sustainability_analysis(
        self,
        values: CellValue | Sequence[CellValue] | Sequence[Sequence[CellValue]],
    ) -> RangeAssignment:
        return _apply_range(
            self, shape=(1, 1), addresses=['START!K10'], values=values
        )

    def set_input_5_local_debt_financing_g00190_by_year(
        self,
        values: Mapping[int, CellValue] | Sequence[CellValue],
        *,
        start_year: int | None = None,
        strict: bool = True,
    ) -> YearRowAssignment:
        if isinstance(values, MappingABC):
            return _apply_year_row_mapping(
                self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043), year_to_addresses={2024: ("'Input 5 - Local-debt Financing'!AE254", "'Input 5 - Local-debt Financing'!AE278", "'Input 5 - Local-debt Financing'!AE302", "'Input 5 - Local-debt Financing'!AG254", "'Input 5 - Local-debt Financing'!AG278", "'Input 5 - Local-debt Financing'!AG302", "'Input 5 - Local-debt Financing'!AG468", "'Input 5 - Local-debt Financing'!AG492", "'Input 5 - Local-debt Financing'!AH254", "'Input 5 - Local-debt Financing'!AH278", "'Input 5 - Local-debt Financing'!AH302", "'Input 5 - Local-debt Financing'!AH468", "'Input 5 - Local-debt Financing'!AH492", "'Input 5 - Local-debt Financing'!AI254", "'Input 5 - Local-debt Financing'!AI278", "'Input 5 - Local-debt Financing'!AI302", "'Input 5 - Local-debt Financing'!AI468", "'Input 5 - Local-debt Financing'!AI492", "'Input 5 - Local-debt Financing'!AJ254", "'Input 5 - Local-debt Financing'!AJ278", "'Input 5 - Local-debt Financing'!AJ302", "'Input 5 - Local-debt Financing'!AJ468", "'Input 5 - Local-debt Financing'!AJ492", "'Input 5 - Local-debt Financing'!AK254", "'Input 5 - Local-debt Financing'!AK278", "'Input 5 - Local-debt Financing'!AK302", "'Input 5 - Local-debt Financing'!AK468", "'Input 5 - Local-debt Financing'!AK492", "'Input 5 - Local-debt Financing'!AL254", "'Input 5 - Local-debt Financing'!AL278", "'Input 5 - Local-debt Financing'!AL302", "'Input 5 - Local-debt Financing'!AL468", "'Input 5 - Local-debt Financing'!AL492", "'Input 5 - Local-debt Financing'!AM254", "'Input 5 - Local-debt Financing'!AM278", "'Input 5 - Local-debt Financing'!AM302", "'Input 5 - Local-debt Financing'!AM468", "'Input 5 - Local-debt Financing'!AM492", "'Input 5 - Local-debt Financing'!AN254", "'Input 5 - Local-debt Financing'!AN278", "'Input 5 - Local-debt Financing'!AN302", "'Input 5 - Local-debt Financing'!AN468", "'Input 5 - Local-debt Financing'!AN492", "'Input 5 - Local-debt Financing'!AO254", "'Input 5 - Local-debt Financing'!AO278", "'Input 5 - Local-debt Financing'!AO302", "'Input 5 - Local-debt Financing'!AO468", "'Input 5 - Local-debt Financing'!AO492", "'Input 5 - Local-debt Financing'!AP254", "'Input 5 - Local-debt Financing'!AP278", "'Input 5 - Local-debt Financing'!AP302", "'Input 5 - Local-debt Financing'!AP468", "'Input 5 - Local-debt Financing'!AP492", "'Input 5 - Local-debt Financing'!AQ254", "'Input 5 - Local-debt Financing'!AQ278", "'Input 5 - Local-debt Financing'!AQ302", "'Input 5 - Local-debt Financing'!AQ468", "'Input 5 - Local-debt Financing'!AQ492", "'Input 5 - Local-debt Financing'!AR254", "'Input 5 - Local-debt Financing'!AR278", "'Input 5 - Local-debt Financing'!AR302", "'Input 5 - Local-debt Financing'!AR468", "'Input 5 - Local-debt Financing'!AR492", "'Input 5 - Local-debt Financing'!AS254", "'Input 5 - Local-debt Financing'!AS278", "'Input 5 - Local-debt Financing'!AS302", "'Input 5 - Local-debt Financing'!AS468", "'Input 5 - Local-debt Financing'!AS492", "'Input 5 - Local-debt Financing'!AT254", "'Input 5 - Local-debt Financing'!AT278", "'Input 5 - Local-debt Financing'!AT302", "'Input 5 - Local-debt Financing'!AT468", "'Input 5 - Local-debt Financing'!AT492", "'Input 5 - Local-debt Financing'!AU254", "'Input 5 - Local-debt Financing'!AU278", "'Input 5 - Local-debt Financing'!AU302", "'Input 5 - Local-debt Financing'!AU468", "'Input 5 - Local-debt Financing'!AU492", "'Input 5 - Local-debt Financing'!AV254", "'Input 5 - Local-debt Financing'!AV278", "'Input 5 - Local-debt Financing'!AV302", "'Input 5 - Local-debt Financing'!AV468", "'Input 5 - Local-debt Financing'!AV492", "'Input 5 - Local-debt Financing'!AW254", "'Input 5 - Local-debt Financing'!AW278", "'Input 5 - Local-debt Financing'!AW302", "'Input 5 - Local-debt Financing'!AW468", "'Input 5 - Local-debt Financing'!AW492", "'Input 5 - Local-debt Financing'!AX254", "'Input 5 - Local-debt Financing'!AX278", "'Input 5 - Local-debt Financing'!AX302", "'Input 5 - Local-debt Financing'!AX468", "'Input 5 - Local-debt Financing'!AX492", "'Input 5 - Local-debt Financing'!AY254", "'Input 5 - Local-debt Financing'!AY278", "'Input 5 - Local-debt Financing'!AY302", "'Input 5 - Local-debt Financing'!AY468", "'Input 5 - Local-debt Financing'!AY492"), 2025: ("'Input 5 - Local-debt Financing'!AF255", "'Input 5 - Local-debt Financing'!AF279", "'Input 5 - Local-debt Financing'!AF303", "'Input 5 - Local-debt Financing'!AF469", "'Input 5 - Local-debt Financing'!AF493", "'Input 5 - Local-debt Financing'!AH255", "'Input 5 - Local-debt Financing'!AH279", "'Input 5 - Local-debt Financing'!AH303", "'Input 5 - Local-debt Financing'!AH469", "'Input 5 - Local-debt Financing'!AH493", "'Input 5 - Local-debt Financing'!AI255", "'Input 5 - Local-debt Financing'!AI279", "'Input 5 - Local-debt Financing'!AI303", "'Input 5 - Local-debt Financing'!AI469", "'Input 5 - Local-debt Financing'!AI493", "'Input 5 - Local-debt Financing'!AJ255", "'Input 5 - Local-debt Financing'!AJ279", "'Input 5 - Local-debt Financing'!AJ303", "'Input 5 - Local-debt Financing'!AJ469", "'Input 5 - Local-debt Financing'!AJ493", "'Input 5 - Local-debt Financing'!AK255", "'Input 5 - Local-debt Financing'!AK279", "'Input 5 - Local-debt Financing'!AK303", "'Input 5 - Local-debt Financing'!AK469", "'Input 5 - Local-debt Financing'!AK493", "'Input 5 - Local-debt Financing'!AL255", "'Input 5 - Local-debt Financing'!AL279", "'Input 5 - Local-debt Financing'!AL303", "'Input 5 - Local-debt Financing'!AL469", "'Input 5 - Local-debt Financing'!AL493", "'Input 5 - Local-debt Financing'!AM255", "'Input 5 - Local-debt Financing'!AM279", "'Input 5 - Local-debt Financing'!AM303", "'Input 5 - Local-debt Financing'!AM469", "'Input 5 - Local-debt Financing'!AM493", "'Input 5 - Local-debt Financing'!AN255", "'Input 5 - Local-debt Financing'!AN279", "'Input 5 - Local-debt Financing'!AN303", "'Input 5 - Local-debt Financing'!AN469", "'Input 5 - Local-debt Financing'!AN493", "'Input 5 - Local-debt Financing'!AO255", "'Input 5 - Local-debt Financing'!AO279", "'Input 5 - Local-debt Financing'!AO303", "'Input 5 - Local-debt Financing'!AO469", "'Input 5 - Local-debt Financing'!AO493", "'Input 5 - Local-debt Financing'!AP255", "'Input 5 - Local-debt Financing'!AP279", "'Input 5 - Local-debt Financing'!AP303", "'Input 5 - Local-debt Financing'!AP469", "'Input 5 - Local-debt Financing'!AP493", "'Input 5 - Local-debt Financing'!AQ255", "'Input 5 - Local-debt Financing'!AQ279", "'Input 5 - Local-debt Financing'!AQ303", "'Input 5 - Local-debt Financing'!AQ469", "'Input 5 - Local-debt Financing'!AQ493", "'Input 5 - Local-debt Financing'!AR255", "'Input 5 - Local-debt Financing'!AR279", "'Input 5 - Local-debt Financing'!AR303", "'Input 5 - Local-debt Financing'!AR469", "'Input 5 - Local-debt Financing'!AR493", "'Input 5 - Local-debt Financing'!AS255", "'Input 5 - Local-debt Financing'!AS279", "'Input 5 - Local-debt Financing'!AS303", "'Input 5 - Local-debt Financing'!AS469", "'Input 5 - Local-debt Financing'!AS493", "'Input 5 - Local-debt Financing'!AT255", "'Input 5 - Local-debt Financing'!AT279", "'Input 5 - Local-debt Financing'!AT303", "'Input 5 - Local-debt Financing'!AT469", "'Input 5 - Local-debt Financing'!AT493", "'Input 5 - Local-debt Financing'!AU255", "'Input 5 - Local-debt Financing'!AU279", "'Input 5 - Local-debt Financing'!AU303", "'Input 5 - Local-debt Financing'!AU469", "'Input 5 - Local-debt Financing'!AU493", "'Input 5 - Local-debt Financing'!AV255", "'Input 5 - Local-debt Financing'!AV279", "'Input 5 - Local-debt Financing'!AV303", "'Input 5 - Local-debt Financing'!AV469", "'Input 5 - Local-debt Financing'!AV493", "'Input 5 - Local-debt Financing'!AW255", "'Input 5 - Local-debt Financing'!AW279", "'Input 5 - Local-debt Financing'!AW303", "'Input 5 - Local-debt Financing'!AW469", "'Input 5 - Local-debt Financing'!AW493", "'Input 5 - Local-debt Financing'!AX255", "'Input 5 - Local-debt Financing'!AX279", "'Input 5 - Local-debt Financing'!AX303", "'Input 5 - Local-debt Financing'!AX469", "'Input 5 - Local-debt Financing'!AX493", "'Input 5 - Local-debt Financing'!AY255", "'Input 5 - Local-debt Financing'!AY279", "'Input 5 - Local-debt Financing'!AY303", "'Input 5 - Local-debt Financing'!AY469", "'Input 5 - Local-debt Financing'!AY493"), 2026: ("'Input 5 - Local-debt Financing'!AG256", "'Input 5 - Local-debt Financing'!AG280", "'Input 5 - Local-debt Financing'!AG304", "'Input 5 - Local-debt Financing'!AG470", "'Input 5 - Local-debt Financing'!AG494", "'Input 5 - Local-debt Financing'!AI256", "'Input 5 - Local-debt Financing'!AI280", "'Input 5 - Local-debt Financing'!AI304", "'Input 5 - Local-debt Financing'!AI470", "'Input 5 - Local-debt Financing'!AI494", "'Input 5 - Local-debt Financing'!AJ256", "'Input 5 - Local-debt Financing'!AJ280", "'Input 5 - Local-debt Financing'!AJ304", "'Input 5 - Local-debt Financing'!AJ470", "'Input 5 - Local-debt Financing'!AJ494", "'Input 5 - Local-debt Financing'!AK256", "'Input 5 - Local-debt Financing'!AK280", "'Input 5 - Local-debt Financing'!AK304", "'Input 5 - Local-debt Financing'!AK470", "'Input 5 - Local-debt Financing'!AK494", "'Input 5 - Local-debt Financing'!AL256", "'Input 5 - Local-debt Financing'!AL280", "'Input 5 - Local-debt Financing'!AL304", "'Input 5 - Local-debt Financing'!AL470", "'Input 5 - Local-debt Financing'!AL494", "'Input 5 - Local-debt Financing'!AM256", "'Input 5 - Local-debt Financing'!AM280", "'Input 5 - Local-debt Financing'!AM304", "'Input 5 - Local-debt Financing'!AM470", "'Input 5 - Local-debt Financing'!AM494", "'Input 5 - Local-debt Financing'!AN256", "'Input 5 - Local-debt Financing'!AN280", "'Input 5 - Local-debt Financing'!AN304", "'Input 5 - Local-debt Financing'!AN470", "'Input 5 - Local-debt Financing'!AN494", "'Input 5 - Local-debt Financing'!AO256", "'Input 5 - Local-debt Financing'!AO280", "'Input 5 - Local-debt Financing'!AO304", "'Input 5 - Local-debt Financing'!AO470", "'Input 5 - Local-debt Financing'!AO494", "'Input 5 - Local-debt Financing'!AP256", "'Input 5 - Local-debt Financing'!AP280", "'Input 5 - Local-debt Financing'!AP304", "'Input 5 - Local-debt Financing'!AP470", "'Input 5 - Local-debt Financing'!AP494", "'Input 5 - Local-debt Financing'!AQ256", "'Input 5 - Local-debt Financing'!AQ280", "'Input 5 - Local-debt Financing'!AQ304", "'Input 5 - Local-debt Financing'!AQ470", "'Input 5 - Local-debt Financing'!AQ494", "'Input 5 - Local-debt Financing'!AR256", "'Input 5 - Local-debt Financing'!AR280", "'Input 5 - Local-debt Financing'!AR304", "'Input 5 - Local-debt Financing'!AR470", "'Input 5 - Local-debt Financing'!AR494", "'Input 5 - Local-debt Financing'!AS256", "'Input 5 - Local-debt Financing'!AS280", "'Input 5 - Local-debt Financing'!AS304", "'Input 5 - Local-debt Financing'!AS470", "'Input 5 - Local-debt Financing'!AS494", "'Input 5 - Local-debt Financing'!AT256", "'Input 5 - Local-debt Financing'!AT280", "'Input 5 - Local-debt Financing'!AT304", "'Input 5 - Local-debt Financing'!AT470", "'Input 5 - Local-debt Financing'!AT494", "'Input 5 - Local-debt Financing'!AU256", "'Input 5 - Local-debt Financing'!AU280", "'Input 5 - Local-debt Financing'!AU304", "'Input 5 - Local-debt Financing'!AU470", "'Input 5 - Local-debt Financing'!AU494", "'Input 5 - Local-debt Financing'!AV256", "'Input 5 - Local-debt Financing'!AV280", "'Input 5 - Local-debt Financing'!AV304", "'Input 5 - Local-debt Financing'!AV470", "'Input 5 - Local-debt Financing'!AV494", "'Input 5 - Local-debt Financing'!AW256", "'Input 5 - Local-debt Financing'!AW280", "'Input 5 - Local-debt Financing'!AW304", "'Input 5 - Local-debt Financing'!AW470", "'Input 5 - Local-debt Financing'!AW494", "'Input 5 - Local-debt Financing'!AX256", "'Input 5 - Local-debt Financing'!AX280", "'Input 5 - Local-debt Financing'!AX304", "'Input 5 - Local-debt Financing'!AX470", "'Input 5 - Local-debt Financing'!AX494", "'Input 5 - Local-debt Financing'!AY256", "'Input 5 - Local-debt Financing'!AY280", "'Input 5 - Local-debt Financing'!AY304", "'Input 5 - Local-debt Financing'!AY470", "'Input 5 - Local-debt Financing'!AY494"), 2027: ("'Input 5 - Local-debt Financing'!AH257", "'Input 5 - Local-debt Financing'!AH281", "'Input 5 - Local-debt Financing'!AH305", "'Input 5 - Local-debt Financing'!AH471", "'Input 5 - Local-debt Financing'!AH495", "'Input 5 - Local-debt Financing'!AJ257", "'Input 5 - Local-debt Financing'!AJ281", "'Input 5 - Local-debt Financing'!AJ305", "'Input 5 - Local-debt Financing'!AJ471", "'Input 5 - Local-debt Financing'!AJ495", "'Input 5 - Local-debt Financing'!AK257", "'Input 5 - Local-debt Financing'!AK281", "'Input 5 - Local-debt Financing'!AK305", "'Input 5 - Local-debt Financing'!AK471", "'Input 5 - Local-debt Financing'!AK495", "'Input 5 - Local-debt Financing'!AL257", "'Input 5 - Local-debt Financing'!AL281", "'Input 5 - Local-debt Financing'!AL305", "'Input 5 - Local-debt Financing'!AL471", "'Input 5 - Local-debt Financing'!AL495", "'Input 5 - Local-debt Financing'!AM257", "'Input 5 - Local-debt Financing'!AM281", "'Input 5 - Local-debt Financing'!AM305", "'Input 5 - Local-debt Financing'!AM471", "'Input 5 - Local-debt Financing'!AM495", "'Input 5 - Local-debt Financing'!AN257", "'Input 5 - Local-debt Financing'!AN281", "'Input 5 - Local-debt Financing'!AN305", "'Input 5 - Local-debt Financing'!AN471", "'Input 5 - Local-debt Financing'!AN495", "'Input 5 - Local-debt Financing'!AO257", "'Input 5 - Local-debt Financing'!AO281", "'Input 5 - Local-debt Financing'!AO305", "'Input 5 - Local-debt Financing'!AO471", "'Input 5 - Local-debt Financing'!AO495", "'Input 5 - Local-debt Financing'!AP257", "'Input 5 - Local-debt Financing'!AP281", "'Input 5 - Local-debt Financing'!AP305", "'Input 5 - Local-debt Financing'!AP471", "'Input 5 - Local-debt Financing'!AP495", "'Input 5 - Local-debt Financing'!AQ257", "'Input 5 - Local-debt Financing'!AQ281", "'Input 5 - Local-debt Financing'!AQ305", "'Input 5 - Local-debt Financing'!AQ471", "'Input 5 - Local-debt Financing'!AQ495", "'Input 5 - Local-debt Financing'!AR257", "'Input 5 - Local-debt Financing'!AR281", "'Input 5 - Local-debt Financing'!AR305", "'Input 5 - Local-debt Financing'!AR471", "'Input 5 - Local-debt Financing'!AR495", "'Input 5 - Local-debt Financing'!AS257", "'Input 5 - Local-debt Financing'!AS281", "'Input 5 - Local-debt Financing'!AS305", "'Input 5 - Local-debt Financing'!AS471", "'Input 5 - Local-debt Financing'!AS495", "'Input 5 - Local-debt Financing'!AT257", "'Input 5 - Local-debt Financing'!AT281", "'Input 5 - Local-debt Financing'!AT305", "'Input 5 - Local-debt Financing'!AT471", "'Input 5 - Local-debt Financing'!AT495", "'Input 5 - Local-debt Financing'!AU257", "'Input 5 - Local-debt Financing'!AU281", "'Input 5 - Local-debt Financing'!AU305", "'Input 5 - Local-debt Financing'!AU471", "'Input 5 - Local-debt Financing'!AU495", "'Input 5 - Local-debt Financing'!AV257", "'Input 5 - Local-debt Financing'!AV281", "'Input 5 - Local-debt Financing'!AV305", "'Input 5 - Local-debt Financing'!AV471", "'Input 5 - Local-debt Financing'!AV495", "'Input 5 - Local-debt Financing'!AW257", "'Input 5 - Local-debt Financing'!AW281", "'Input 5 - Local-debt Financing'!AW305", "'Input 5 - Local-debt Financing'!AW471", "'Input 5 - Local-debt Financing'!AW495", "'Input 5 - Local-debt Financing'!AX257", "'Input 5 - Local-debt Financing'!AX281", "'Input 5 - Local-debt Financing'!AX305", "'Input 5 - Local-debt Financing'!AX471", "'Input 5 - Local-debt Financing'!AX495", "'Input 5 - Local-debt Financing'!AY257", "'Input 5 - Local-debt Financing'!AY281", "'Input 5 - Local-debt Financing'!AY305", "'Input 5 - Local-debt Financing'!AY471", "'Input 5 - Local-debt Financing'!AY495"), 2028: ("'Input 5 - Local-debt Financing'!AI258", "'Input 5 - Local-debt Financing'!AI282", "'Input 5 - Local-debt Financing'!AI306", "'Input 5 - Local-debt Financing'!AI472", "'Input 5 - Local-debt Financing'!AI496", "'Input 5 - Local-debt Financing'!AK282", "'Input 5 - Local-debt Financing'!AK496", "'Input 5 - Local-debt Financing'!AL282", "'Input 5 - Local-debt Financing'!AL496", "'Input 5 - Local-debt Financing'!AM282", "'Input 5 - Local-debt Financing'!AM496", "'Input 5 - Local-debt Financing'!AN282", "'Input 5 - Local-debt Financing'!AN496", "'Input 5 - Local-debt Financing'!AO282", "'Input 5 - Local-debt Financing'!AO496", "'Input 5 - Local-debt Financing'!AP282", "'Input 5 - Local-debt Financing'!AP496", "'Input 5 - Local-debt Financing'!AQ282", "'Input 5 - Local-debt Financing'!AQ496", "'Input 5 - Local-debt Financing'!AR282", "'Input 5 - Local-debt Financing'!AR496", "'Input 5 - Local-debt Financing'!AS282", "'Input 5 - Local-debt Financing'!AS496", "'Input 5 - Local-debt Financing'!AT282", "'Input 5 - Local-debt Financing'!AT496", "'Input 5 - Local-debt Financing'!AU282", "'Input 5 - Local-debt Financing'!AU496", "'Input 5 - Local-debt Financing'!AV282", "'Input 5 - Local-debt Financing'!AV496", "'Input 5 - Local-debt Financing'!AW282", "'Input 5 - Local-debt Financing'!AW496", "'Input 5 - Local-debt Financing'!AX282", "'Input 5 - Local-debt Financing'!AX496", "'Input 5 - Local-debt Financing'!AY282", "'Input 5 - Local-debt Financing'!AY496"), 2029: ("'Input 5 - Local-debt Financing'!AJ259", "'Input 5 - Local-debt Financing'!AJ283", "'Input 5 - Local-debt Financing'!AJ307", "'Input 5 - Local-debt Financing'!AJ473", "'Input 5 - Local-debt Financing'!AJ497", "'Input 5 - Local-debt Financing'!AL283", "'Input 5 - Local-debt Financing'!AL497", "'Input 5 - Local-debt Financing'!AM283", "'Input 5 - Local-debt Financing'!AM497", "'Input 5 - Local-debt Financing'!AN283", "'Input 5 - Local-debt Financing'!AN497", "'Input 5 - Local-debt Financing'!AO283", "'Input 5 - Local-debt Financing'!AO497", "'Input 5 - Local-debt Financing'!AP283", "'Input 5 - Local-debt Financing'!AP497", "'Input 5 - Local-debt Financing'!AQ283", "'Input 5 - Local-debt Financing'!AQ497", "'Input 5 - Local-debt Financing'!AR283", "'Input 5 - Local-debt Financing'!AR497", "'Input 5 - Local-debt Financing'!AS283", "'Input 5 - Local-debt Financing'!AS497", "'Input 5 - Local-debt Financing'!AT283", "'Input 5 - Local-debt Financing'!AT497", "'Input 5 - Local-debt Financing'!AU283", "'Input 5 - Local-debt Financing'!AU497", "'Input 5 - Local-debt Financing'!AV283", "'Input 5 - Local-debt Financing'!AV497", "'Input 5 - Local-debt Financing'!AW283", "'Input 5 - Local-debt Financing'!AW497", "'Input 5 - Local-debt Financing'!AX283", "'Input 5 - Local-debt Financing'!AX497", "'Input 5 - Local-debt Financing'!AY283", "'Input 5 - Local-debt Financing'!AY497"), 2030: ("'Input 5 - Local-debt Financing'!AK260", "'Input 5 - Local-debt Financing'!AK308", "'Input 5 - Local-debt Financing'!AK474", "'Input 5 - Local-debt Financing'!AM284", "'Input 5 - Local-debt Financing'!AM498", "'Input 5 - Local-debt Financing'!AN284", "'Input 5 - Local-debt Financing'!AN498", "'Input 5 - Local-debt Financing'!AO284", "'Input 5 - Local-debt Financing'!AO498", "'Input 5 - Local-debt Financing'!AP284", "'Input 5 - Local-debt Financing'!AP498", "'Input 5 - Local-debt Financing'!AQ284", "'Input 5 - Local-debt Financing'!AQ498", "'Input 5 - Local-debt Financing'!AR284", "'Input 5 - Local-debt Financing'!AR498", "'Input 5 - Local-debt Financing'!AS284", "'Input 5 - Local-debt Financing'!AS498", "'Input 5 - Local-debt Financing'!AT284", "'Input 5 - Local-debt Financing'!AT498", "'Input 5 - Local-debt Financing'!AU284", "'Input 5 - Local-debt Financing'!AU498", "'Input 5 - Local-debt Financing'!AV284", "'Input 5 - Local-debt Financing'!AV498", "'Input 5 - Local-debt Financing'!AW284", "'Input 5 - Local-debt Financing'!AW498", "'Input 5 - Local-debt Financing'!AX284", "'Input 5 - Local-debt Financing'!AX498", "'Input 5 - Local-debt Financing'!AY284", "'Input 5 - Local-debt Financing'!AY498"), 2031: ("'Input 5 - Local-debt Financing'!AL261", "'Input 5 - Local-debt Financing'!AL309", "'Input 5 - Local-debt Financing'!AL475", "'Input 5 - Local-debt Financing'!AN285", "'Input 5 - Local-debt Financing'!AN499", "'Input 5 - Local-debt Financing'!AO285", "'Input 5 - Local-debt Financing'!AO499", "'Input 5 - Local-debt Financing'!AP285", "'Input 5 - Local-debt Financing'!AP499", "'Input 5 - Local-debt Financing'!AQ285", "'Input 5 - Local-debt Financing'!AQ499", "'Input 5 - Local-debt Financing'!AR285", "'Input 5 - Local-debt Financing'!AR499", "'Input 5 - Local-debt Financing'!AS285", "'Input 5 - Local-debt Financing'!AS499", "'Input 5 - Local-debt Financing'!AT285", "'Input 5 - Local-debt Financing'!AT499", "'Input 5 - Local-debt Financing'!AU285", "'Input 5 - Local-debt Financing'!AU499", "'Input 5 - Local-debt Financing'!AV285", "'Input 5 - Local-debt Financing'!AV499", "'Input 5 - Local-debt Financing'!AW285", "'Input 5 - Local-debt Financing'!AW499", "'Input 5 - Local-debt Financing'!AX285", "'Input 5 - Local-debt Financing'!AX499", "'Input 5 - Local-debt Financing'!AY285", "'Input 5 - Local-debt Financing'!AY499"), 2032: ("'Input 5 - Local-debt Financing'!AM262", "'Input 5 - Local-debt Financing'!AM310", "'Input 5 - Local-debt Financing'!AM476", "'Input 5 - Local-debt Financing'!AO286", "'Input 5 - Local-debt Financing'!AO500", "'Input 5 - Local-debt Financing'!AP286", "'Input 5 - Local-debt Financing'!AP500", "'Input 5 - Local-debt Financing'!AQ286", "'Input 5 - Local-debt Financing'!AQ500", "'Input 5 - Local-debt Financing'!AR286", "'Input 5 - Local-debt Financing'!AR500", "'Input 5 - Local-debt Financing'!AS286", "'Input 5 - Local-debt Financing'!AS500", "'Input 5 - Local-debt Financing'!AT286", "'Input 5 - Local-debt Financing'!AT500", "'Input 5 - Local-debt Financing'!AU286", "'Input 5 - Local-debt Financing'!AU500", "'Input 5 - Local-debt Financing'!AV286", "'Input 5 - Local-debt Financing'!AV500", "'Input 5 - Local-debt Financing'!AW286", "'Input 5 - Local-debt Financing'!AW500", "'Input 5 - Local-debt Financing'!AX286", "'Input 5 - Local-debt Financing'!AX500", "'Input 5 - Local-debt Financing'!AY286", "'Input 5 - Local-debt Financing'!AY500"), 2033: ("'Input 5 - Local-debt Financing'!AN263", "'Input 5 - Local-debt Financing'!AN311", "'Input 5 - Local-debt Financing'!AN477", "'Input 5 - Local-debt Financing'!AP287", "'Input 5 - Local-debt Financing'!AP501", "'Input 5 - Local-debt Financing'!AQ287", "'Input 5 - Local-debt Financing'!AQ501", "'Input 5 - Local-debt Financing'!AR287", "'Input 5 - Local-debt Financing'!AR501", "'Input 5 - Local-debt Financing'!AS287", "'Input 5 - Local-debt Financing'!AS501", "'Input 5 - Local-debt Financing'!AT287", "'Input 5 - Local-debt Financing'!AT501", "'Input 5 - Local-debt Financing'!AU287", "'Input 5 - Local-debt Financing'!AU501", "'Input 5 - Local-debt Financing'!AV287", "'Input 5 - Local-debt Financing'!AV501", "'Input 5 - Local-debt Financing'!AW287", "'Input 5 - Local-debt Financing'!AW501", "'Input 5 - Local-debt Financing'!AX287", "'Input 5 - Local-debt Financing'!AX501", "'Input 5 - Local-debt Financing'!AY287", "'Input 5 - Local-debt Financing'!AY501"), 2034: ("'Input 5 - Local-debt Financing'!AO264", "'Input 5 - Local-debt Financing'!AO312", "'Input 5 - Local-debt Financing'!AO478", "'Input 5 - Local-debt Financing'!AQ288", "'Input 5 - Local-debt Financing'!AQ502", "'Input 5 - Local-debt Financing'!AR288", "'Input 5 - Local-debt Financing'!AR502", "'Input 5 - Local-debt Financing'!AS288", "'Input 5 - Local-debt Financing'!AS502", "'Input 5 - Local-debt Financing'!AT288", "'Input 5 - Local-debt Financing'!AT502", "'Input 5 - Local-debt Financing'!AU288", "'Input 5 - Local-debt Financing'!AU502", "'Input 5 - Local-debt Financing'!AV288", "'Input 5 - Local-debt Financing'!AV502", "'Input 5 - Local-debt Financing'!AW288", "'Input 5 - Local-debt Financing'!AW502", "'Input 5 - Local-debt Financing'!AX288", "'Input 5 - Local-debt Financing'!AX502", "'Input 5 - Local-debt Financing'!AY288", "'Input 5 - Local-debt Financing'!AY502"), 2035: ("'Input 5 - Local-debt Financing'!AP265", "'Input 5 - Local-debt Financing'!AP313", "'Input 5 - Local-debt Financing'!AP479", "'Input 5 - Local-debt Financing'!AR289", "'Input 5 - Local-debt Financing'!AR503", "'Input 5 - Local-debt Financing'!AS289", "'Input 5 - Local-debt Financing'!AS503", "'Input 5 - Local-debt Financing'!AT289", "'Input 5 - Local-debt Financing'!AT503", "'Input 5 - Local-debt Financing'!AU289", "'Input 5 - Local-debt Financing'!AU503", "'Input 5 - Local-debt Financing'!AV289", "'Input 5 - Local-debt Financing'!AV503", "'Input 5 - Local-debt Financing'!AW289", "'Input 5 - Local-debt Financing'!AW503", "'Input 5 - Local-debt Financing'!AX289", "'Input 5 - Local-debt Financing'!AX503", "'Input 5 - Local-debt Financing'!AY289", "'Input 5 - Local-debt Financing'!AY503"), 2036: ("'Input 5 - Local-debt Financing'!AQ266", "'Input 5 - Local-debt Financing'!AQ314", "'Input 5 - Local-debt Financing'!AQ480", "'Input 5 - Local-debt Financing'!AS290", "'Input 5 - Local-debt Financing'!AS504", "'Input 5 - Local-debt Financing'!AT290", "'Input 5 - Local-debt Financing'!AT504", "'Input 5 - Local-debt Financing'!AU290", "'Input 5 - Local-debt Financing'!AU504", "'Input 5 - Local-debt Financing'!AV290", "'Input 5 - Local-debt Financing'!AV504", "'Input 5 - Local-debt Financing'!AW290", "'Input 5 - Local-debt Financing'!AW504", "'Input 5 - Local-debt Financing'!AX290", "'Input 5 - Local-debt Financing'!AX504", "'Input 5 - Local-debt Financing'!AY290", "'Input 5 - Local-debt Financing'!AY504"), 2037: ("'Input 5 - Local-debt Financing'!AR267", "'Input 5 - Local-debt Financing'!AR315", "'Input 5 - Local-debt Financing'!AR481", "'Input 5 - Local-debt Financing'!AT291", "'Input 5 - Local-debt Financing'!AT505", "'Input 5 - Local-debt Financing'!AU291", "'Input 5 - Local-debt Financing'!AU505", "'Input 5 - Local-debt Financing'!AV291", "'Input 5 - Local-debt Financing'!AV505", "'Input 5 - Local-debt Financing'!AW291", "'Input 5 - Local-debt Financing'!AW505", "'Input 5 - Local-debt Financing'!AX291", "'Input 5 - Local-debt Financing'!AX505", "'Input 5 - Local-debt Financing'!AY291", "'Input 5 - Local-debt Financing'!AY505"), 2038: ("'Input 5 - Local-debt Financing'!AS268", "'Input 5 - Local-debt Financing'!AS316", "'Input 5 - Local-debt Financing'!AS482", "'Input 5 - Local-debt Financing'!AU292", "'Input 5 - Local-debt Financing'!AU506", "'Input 5 - Local-debt Financing'!AV292", "'Input 5 - Local-debt Financing'!AV506", "'Input 5 - Local-debt Financing'!AW292", "'Input 5 - Local-debt Financing'!AW506", "'Input 5 - Local-debt Financing'!AX292", "'Input 5 - Local-debt Financing'!AX506", "'Input 5 - Local-debt Financing'!AY292", "'Input 5 - Local-debt Financing'!AY506"), 2039: ("'Input 5 - Local-debt Financing'!AT269", "'Input 5 - Local-debt Financing'!AT317", "'Input 5 - Local-debt Financing'!AT483", "'Input 5 - Local-debt Financing'!AV293", "'Input 5 - Local-debt Financing'!AV507", "'Input 5 - Local-debt Financing'!AW293", "'Input 5 - Local-debt Financing'!AW507", "'Input 5 - Local-debt Financing'!AX293", "'Input 5 - Local-debt Financing'!AX507", "'Input 5 - Local-debt Financing'!AY293", "'Input 5 - Local-debt Financing'!AY507"), 2040: ("'Input 5 - Local-debt Financing'!AU270", "'Input 5 - Local-debt Financing'!AU318", "'Input 5 - Local-debt Financing'!AU484", "'Input 5 - Local-debt Financing'!AW294", "'Input 5 - Local-debt Financing'!AW508", "'Input 5 - Local-debt Financing'!AX294", "'Input 5 - Local-debt Financing'!AX508", "'Input 5 - Local-debt Financing'!AY294", "'Input 5 - Local-debt Financing'!AY508"), 2041: ("'Input 5 - Local-debt Financing'!AV271", "'Input 5 - Local-debt Financing'!AV319", "'Input 5 - Local-debt Financing'!AV485", "'Input 5 - Local-debt Financing'!AX295", "'Input 5 - Local-debt Financing'!AX509", "'Input 5 - Local-debt Financing'!AY295", "'Input 5 - Local-debt Financing'!AY509"), 2042: ("'Input 5 - Local-debt Financing'!AW272", "'Input 5 - Local-debt Financing'!AW320", "'Input 5 - Local-debt Financing'!AW486", "'Input 5 - Local-debt Financing'!AY296", "'Input 5 - Local-debt Financing'!AY510"), 2043: ("'Input 5 - Local-debt Financing'!AX273", "'Input 5 - Local-debt Financing'!AX321", "'Input 5 - Local-debt Financing'!AX487")},
                values_by_year=values, strict=strict,
            )
        if not isinstance(values, SequenceABC):
            raise TypeError("Expected a mapping or sequence for year-row inputs")
        if start_year is None:
            raise TypeError("start_year is required for sequence inputs")
        return _apply_year_row_array(
            self, years=(2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043), year_to_addresses={2024: ("'Input 5 - Local-debt Financing'!AE254", "'Input 5 - Local-debt Financing'!AE278", "'Input 5 - Local-debt Financing'!AE302", "'Input 5 - Local-debt Financing'!AG254", "'Input 5 - Local-debt Financing'!AG278", "'Input 5 - Local-debt Financing'!AG302", "'Input 5 - Local-debt Financing'!AG468", "'Input 5 - Local-debt Financing'!AG492", "'Input 5 - Local-debt Financing'!AH254", "'Input 5 - Local-debt Financing'!AH278", "'Input 5 - Local-debt Financing'!AH302", "'Input 5 - Local-debt Financing'!AH468", "'Input 5 - Local-debt Financing'!AH492", "'Input 5 - Local-debt Financing'!AI254", "'Input 5 - Local-debt Financing'!AI278", "'Input 5 - Local-debt Financing'!AI302", "'Input 5 - Local-debt Financing'!AI468", "'Input 5 - Local-debt Financing'!AI492", "'Input 5 - Local-debt Financing'!AJ254", "'Input 5 - Local-debt Financing'!AJ278", "'Input 5 - Local-debt Financing'!AJ302", "'Input 5 - Local-debt Financing'!AJ468", "'Input 5 - Local-debt Financing'!AJ492", "'Input 5 - Local-debt Financing'!AK254", "'Input 5 - Local-debt Financing'!AK278", "'Input 5 - Local-debt Financing'!AK302", "'Input 5 - Local-debt Financing'!AK468", "'Input 5 - Local-debt Financing'!AK492", "'Input 5 - Local-debt Financing'!AL254", "'Input 5 - Local-debt Financing'!AL278", "'Input 5 - Local-debt Financing'!AL302", "'Input 5 - Local-debt Financing'!AL468", "'Input 5 - Local-debt Financing'!AL492", "'Input 5 - Local-debt Financing'!AM254", "'Input 5 - Local-debt Financing'!AM278", "'Input 5 - Local-debt Financing'!AM302", "'Input 5 - Local-debt Financing'!AM468", "'Input 5 - Local-debt Financing'!AM492", "'Input 5 - Local-debt Financing'!AN254", "'Input 5 - Local-debt Financing'!AN278", "'Input 5 - Local-debt Financing'!AN302", "'Input 5 - Local-debt Financing'!AN468", "'Input 5 - Local-debt Financing'!AN492", "'Input 5 - Local-debt Financing'!AO254", "'Input 5 - Local-debt Financing'!AO278", "'Input 5 - Local-debt Financing'!AO302", "'Input 5 - Local-debt Financing'!AO468", "'Input 5 - Local-debt Financing'!AO492", "'Input 5 - Local-debt Financing'!AP254", "'Input 5 - Local-debt Financing'!AP278", "'Input 5 - Local-debt Financing'!AP302", "'Input 5 - Local-debt Financing'!AP468", "'Input 5 - Local-debt Financing'!AP492", "'Input 5 - Local-debt Financing'!AQ254", "'Input 5 - Local-debt Financing'!AQ278", "'Input 5 - Local-debt Financing'!AQ302", "'Input 5 - Local-debt Financing'!AQ468", "'Input 5 - Local-debt Financing'!AQ492", "'Input 5 - Local-debt Financing'!AR254", "'Input 5 - Local-debt Financing'!AR278", "'Input 5 - Local-debt Financing'!AR302", "'Input 5 - Local-debt Financing'!AR468", "'Input 5 - Local-debt Financing'!AR492", "'Input 5 - Local-debt Financing'!AS254", "'Input 5 - Local-debt Financing'!AS278", "'Input 5 - Local-debt Financing'!AS302", "'Input 5 - Local-debt Financing'!AS468", "'Input 5 - Local-debt Financing'!AS492", "'Input 5 - Local-debt Financing'!AT254", "'Input 5 - Local-debt Financing'!AT278", "'Input 5 - Local-debt Financing'!AT302", "'Input 5 - Local-debt Financing'!AT468", "'Input 5 - Local-debt Financing'!AT492", "'Input 5 - Local-debt Financing'!AU254", "'Input 5 - Local-debt Financing'!AU278", "'Input 5 - Local-debt Financing'!AU302", "'Input 5 - Local-debt Financing'!AU468", "'Input 5 - Local-debt Financing'!AU492", "'Input 5 - Local-debt Financing'!AV254", "'Input 5 - Local-debt Financing'!AV278", "'Input 5 - Local-debt Financing'!AV302", "'Input 5 - Local-debt Financing'!AV468", "'Input 5 - Local-debt Financing'!AV492", "'Input 5 - Local-debt Financing'!AW254", "'Input 5 - Local-debt Financing'!AW278", "'Input 5 - Local-debt Financing'!AW302", "'Input 5 - Local-debt Financing'!AW468", "'Input 5 - Local-debt Financing'!AW492", "'Input 5 - Local-debt Financing'!AX254", "'Input 5 - Local-debt Financing'!AX278", "'Input 5 - Local-debt Financing'!AX302", "'Input 5 - Local-debt Financing'!AX468", "'Input 5 - Local-debt Financing'!AX492", "'Input 5 - Local-debt Financing'!AY254", "'Input 5 - Local-debt Financing'!AY278", "'Input 5 - Local-debt Financing'!AY302", "'Input 5 - Local-debt Financing'!AY468", "'Input 5 - Local-debt Financing'!AY492"), 2025: ("'Input 5 - Local-debt Financing'!AF255", "'Input 5 - Local-debt Financing'!AF279", "'Input 5 - Local-debt Financing'!AF303", "'Input 5 - Local-debt Financing'!AF469", "'Input 5 - Local-debt Financing'!AF493", "'Input 5 - Local-debt Financing'!AH255", "'Input 5 - Local-debt Financing'!AH279", "'Input 5 - Local-debt Financing'!AH303", "'Input 5 - Local-debt Financing'!AH469", "'Input 5 - Local-debt Financing'!AH493", "'Input 5 - Local-debt Financing'!AI255", "'Input 5 - Local-debt Financing'!AI279", "'Input 5 - Local-debt Financing'!AI303", "'Input 5 - Local-debt Financing'!AI469", "'Input 5 - Local-debt Financing'!AI493", "'Input 5 - Local-debt Financing'!AJ255", "'Input 5 - Local-debt Financing'!AJ279", "'Input 5 - Local-debt Financing'!AJ303", "'Input 5 - Local-debt Financing'!AJ469", "'Input 5 - Local-debt Financing'!AJ493", "'Input 5 - Local-debt Financing'!AK255", "'Input 5 - Local-debt Financing'!AK279", "'Input 5 - Local-debt Financing'!AK303", "'Input 5 - Local-debt Financing'!AK469", "'Input 5 - Local-debt Financing'!AK493", "'Input 5 - Local-debt Financing'!AL255", "'Input 5 - Local-debt Financing'!AL279", "'Input 5 - Local-debt Financing'!AL303", "'Input 5 - Local-debt Financing'!AL469", "'Input 5 - Local-debt Financing'!AL493", "'Input 5 - Local-debt Financing'!AM255", "'Input 5 - Local-debt Financing'!AM279", "'Input 5 - Local-debt Financing'!AM303", "'Input 5 - Local-debt Financing'!AM469", "'Input 5 - Local-debt Financing'!AM493", "'Input 5 - Local-debt Financing'!AN255", "'Input 5 - Local-debt Financing'!AN279", "'Input 5 - Local-debt Financing'!AN303", "'Input 5 - Local-debt Financing'!AN469", "'Input 5 - Local-debt Financing'!AN493", "'Input 5 - Local-debt Financing'!AO255", "'Input 5 - Local-debt Financing'!AO279", "'Input 5 - Local-debt Financing'!AO303", "'Input 5 - Local-debt Financing'!AO469", "'Input 5 - Local-debt Financing'!AO493", "'Input 5 - Local-debt Financing'!AP255", "'Input 5 - Local-debt Financing'!AP279", "'Input 5 - Local-debt Financing'!AP303", "'Input 5 - Local-debt Financing'!AP469", "'Input 5 - Local-debt Financing'!AP493", "'Input 5 - Local-debt Financing'!AQ255", "'Input 5 - Local-debt Financing'!AQ279", "'Input 5 - Local-debt Financing'!AQ303", "'Input 5 - Local-debt Financing'!AQ469", "'Input 5 - Local-debt Financing'!AQ493", "'Input 5 - Local-debt Financing'!AR255", "'Input 5 - Local-debt Financing'!AR279", "'Input 5 - Local-debt Financing'!AR303", "'Input 5 - Local-debt Financing'!AR469", "'Input 5 - Local-debt Financing'!AR493", "'Input 5 - Local-debt Financing'!AS255", "'Input 5 - Local-debt Financing'!AS279", "'Input 5 - Local-debt Financing'!AS303", "'Input 5 - Local-debt Financing'!AS469", "'Input 5 - Local-debt Financing'!AS493", "'Input 5 - Local-debt Financing'!AT255", "'Input 5 - Local-debt Financing'!AT279", "'Input 5 - Local-debt Financing'!AT303", "'Input 5 - Local-debt Financing'!AT469", "'Input 5 - Local-debt Financing'!AT493", "'Input 5 - Local-debt Financing'!AU255", "'Input 5 - Local-debt Financing'!AU279", "'Input 5 - Local-debt Financing'!AU303", "'Input 5 - Local-debt Financing'!AU469", "'Input 5 - Local-debt Financing'!AU493", "'Input 5 - Local-debt Financing'!AV255", "'Input 5 - Local-debt Financing'!AV279", "'Input 5 - Local-debt Financing'!AV303", "'Input 5 - Local-debt Financing'!AV469", "'Input 5 - Local-debt Financing'!AV493", "'Input 5 - Local-debt Financing'!AW255", "'Input 5 - Local-debt Financing'!AW279", "'Input 5 - Local-debt Financing'!AW303", "'Input 5 - Local-debt Financing'!AW469", "'Input 5 - Local-debt Financing'!AW493", "'Input 5 - Local-debt Financing'!AX255", "'Input 5 - Local-debt Financing'!AX279", "'Input 5 - Local-debt Financing'!AX303", "'Input 5 - Local-debt Financing'!AX469", "'Input 5 - Local-debt Financing'!AX493", "'Input 5 - Local-debt Financing'!AY255", "'Input 5 - Local-debt Financing'!AY279", "'Input 5 - Local-debt Financing'!AY303", "'Input 5 - Local-debt Financing'!AY469", "'Input 5 - Local-debt Financing'!AY493"), 2026: ("'Input 5 - Local-debt Financing'!AG256", "'Input 5 - Local-debt Financing'!AG280", "'Input 5 - Local-debt Financing'!AG304", "'Input 5 - Local-debt Financing'!AG470", "'Input 5 - Local-debt Financing'!AG494", "'Input 5 - Local-debt Financing'!AI256", "'Input 5 - Local-debt Financing'!AI280", "'Input 5 - Local-debt Financing'!AI304", "'Input 5 - Local-debt Financing'!AI470", "'Input 5 - Local-debt Financing'!AI494", "'Input 5 - Local-debt Financing'!AJ256", "'Input 5 - Local-debt Financing'!AJ280", "'Input 5 - Local-debt Financing'!AJ304", "'Input 5 - Local-debt Financing'!AJ470", "'Input 5 - Local-debt Financing'!AJ494", "'Input 5 - Local-debt Financing'!AK256", "'Input 5 - Local-debt Financing'!AK280", "'Input 5 - Local-debt Financing'!AK304", "'Input 5 - Local-debt Financing'!AK470", "'Input 5 - Local-debt Financing'!AK494", "'Input 5 - Local-debt Financing'!AL256", "'Input 5 - Local-debt Financing'!AL280", "'Input 5 - Local-debt Financing'!AL304", "'Input 5 - Local-debt Financing'!AL470", "'Input 5 - Local-debt Financing'!AL494", "'Input 5 - Local-debt Financing'!AM256", "'Input 5 - Local-debt Financing'!AM280", "'Input 5 - Local-debt Financing'!AM304", "'Input 5 - Local-debt Financing'!AM470", "'Input 5 - Local-debt Financing'!AM494", "'Input 5 - Local-debt Financing'!AN256", "'Input 5 - Local-debt Financing'!AN280", "'Input 5 - Local-debt Financing'!AN304", "'Input 5 - Local-debt Financing'!AN470", "'Input 5 - Local-debt Financing'!AN494", "'Input 5 - Local-debt Financing'!AO256", "'Input 5 - Local-debt Financing'!AO280", "'Input 5 - Local-debt Financing'!AO304", "'Input 5 - Local-debt Financing'!AO470", "'Input 5 - Local-debt Financing'!AO494", "'Input 5 - Local-debt Financing'!AP256", "'Input 5 - Local-debt Financing'!AP280", "'Input 5 - Local-debt Financing'!AP304", "'Input 5 - Local-debt Financing'!AP470", "'Input 5 - Local-debt Financing'!AP494", "'Input 5 - Local-debt Financing'!AQ256", "'Input 5 - Local-debt Financing'!AQ280", "'Input 5 - Local-debt Financing'!AQ304", "'Input 5 - Local-debt Financing'!AQ470", "'Input 5 - Local-debt Financing'!AQ494", "'Input 5 - Local-debt Financing'!AR256", "'Input 5 - Local-debt Financing'!AR280", "'Input 5 - Local-debt Financing'!AR304", "'Input 5 - Local-debt Financing'!AR470", "'Input 5 - Local-debt Financing'!AR494", "'Input 5 - Local-debt Financing'!AS256", "'Input 5 - Local-debt Financing'!AS280", "'Input 5 - Local-debt Financing'!AS304", "'Input 5 - Local-debt Financing'!AS470", "'Input 5 - Local-debt Financing'!AS494", "'Input 5 - Local-debt Financing'!AT256", "'Input 5 - Local-debt Financing'!AT280", "'Input 5 - Local-debt Financing'!AT304", "'Input 5 - Local-debt Financing'!AT470", "'Input 5 - Local-debt Financing'!AT494", "'Input 5 - Local-debt Financing'!AU256", "'Input 5 - Local-debt Financing'!AU280", "'Input 5 - Local-debt Financing'!AU304", "'Input 5 - Local-debt Financing'!AU470", "'Input 5 - Local-debt Financing'!AU494", "'Input 5 - Local-debt Financing'!AV256", "'Input 5 - Local-debt Financing'!AV280", "'Input 5 - Local-debt Financing'!AV304", "'Input 5 - Local-debt Financing'!AV470", "'Input 5 - Local-debt Financing'!AV494", "'Input 5 - Local-debt Financing'!AW256", "'Input 5 - Local-debt Financing'!AW280", "'Input 5 - Local-debt Financing'!AW304", "'Input 5 - Local-debt Financing'!AW470", "'Input 5 - Local-debt Financing'!AW494", "'Input 5 - Local-debt Financing'!AX256", "'Input 5 - Local-debt Financing'!AX280", "'Input 5 - Local-debt Financing'!AX304", "'Input 5 - Local-debt Financing'!AX470", "'Input 5 - Local-debt Financing'!AX494", "'Input 5 - Local-debt Financing'!AY256", "'Input 5 - Local-debt Financing'!AY280", "'Input 5 - Local-debt Financing'!AY304", "'Input 5 - Local-debt Financing'!AY470", "'Input 5 - Local-debt Financing'!AY494"), 2027: ("'Input 5 - Local-debt Financing'!AH257", "'Input 5 - Local-debt Financing'!AH281", "'Input 5 - Local-debt Financing'!AH305", "'Input 5 - Local-debt Financing'!AH471", "'Input 5 - Local-debt Financing'!AH495", "'Input 5 - Local-debt Financing'!AJ257", "'Input 5 - Local-debt Financing'!AJ281", "'Input 5 - Local-debt Financing'!AJ305", "'Input 5 - Local-debt Financing'!AJ471", "'Input 5 - Local-debt Financing'!AJ495", "'Input 5 - Local-debt Financing'!AK257", "'Input 5 - Local-debt Financing'!AK281", "'Input 5 - Local-debt Financing'!AK305", "'Input 5 - Local-debt Financing'!AK471", "'Input 5 - Local-debt Financing'!AK495", "'Input 5 - Local-debt Financing'!AL257", "'Input 5 - Local-debt Financing'!AL281", "'Input 5 - Local-debt Financing'!AL305", "'Input 5 - Local-debt Financing'!AL471", "'Input 5 - Local-debt Financing'!AL495", "'Input 5 - Local-debt Financing'!AM257", "'Input 5 - Local-debt Financing'!AM281", "'Input 5 - Local-debt Financing'!AM305", "'Input 5 - Local-debt Financing'!AM471", "'Input 5 - Local-debt Financing'!AM495", "'Input 5 - Local-debt Financing'!AN257", "'Input 5 - Local-debt Financing'!AN281", "'Input 5 - Local-debt Financing'!AN305", "'Input 5 - Local-debt Financing'!AN471", "'Input 5 - Local-debt Financing'!AN495", "'Input 5 - Local-debt Financing'!AO257", "'Input 5 - Local-debt Financing'!AO281", "'Input 5 - Local-debt Financing'!AO305", "'Input 5 - Local-debt Financing'!AO471", "'Input 5 - Local-debt Financing'!AO495", "'Input 5 - Local-debt Financing'!AP257", "'Input 5 - Local-debt Financing'!AP281", "'Input 5 - Local-debt Financing'!AP305", "'Input 5 - Local-debt Financing'!AP471", "'Input 5 - Local-debt Financing'!AP495", "'Input 5 - Local-debt Financing'!AQ257", "'Input 5 - Local-debt Financing'!AQ281", "'Input 5 - Local-debt Financing'!AQ305", "'Input 5 - Local-debt Financing'!AQ471", "'Input 5 - Local-debt Financing'!AQ495", "'Input 5 - Local-debt Financing'!AR257", "'Input 5 - Local-debt Financing'!AR281", "'Input 5 - Local-debt Financing'!AR305", "'Input 5 - Local-debt Financing'!AR471", "'Input 5 - Local-debt Financing'!AR495", "'Input 5 - Local-debt Financing'!AS257", "'Input 5 - Local-debt Financing'!AS281", "'Input 5 - Local-debt Financing'!AS305", "'Input 5 - Local-debt Financing'!AS471", "'Input 5 - Local-debt Financing'!AS495", "'Input 5 - Local-debt Financing'!AT257", "'Input 5 - Local-debt Financing'!AT281", "'Input 5 - Local-debt Financing'!AT305", "'Input 5 - Local-debt Financing'!AT471", "'Input 5 - Local-debt Financing'!AT495", "'Input 5 - Local-debt Financing'!AU257", "'Input 5 - Local-debt Financing'!AU281", "'Input 5 - Local-debt Financing'!AU305", "'Input 5 - Local-debt Financing'!AU471", "'Input 5 - Local-debt Financing'!AU495", "'Input 5 - Local-debt Financing'!AV257", "'Input 5 - Local-debt Financing'!AV281", "'Input 5 - Local-debt Financing'!AV305", "'Input 5 - Local-debt Financing'!AV471", "'Input 5 - Local-debt Financing'!AV495", "'Input 5 - Local-debt Financing'!AW257", "'Input 5 - Local-debt Financing'!AW281", "'Input 5 - Local-debt Financing'!AW305", "'Input 5 - Local-debt Financing'!AW471", "'Input 5 - Local-debt Financing'!AW495", "'Input 5 - Local-debt Financing'!AX257", "'Input 5 - Local-debt Financing'!AX281", "'Input 5 - Local-debt Financing'!AX305", "'Input 5 - Local-debt Financing'!AX471", "'Input 5 - Local-debt Financing'!AX495", "'Input 5 - Local-debt Financing'!AY257", "'Input 5 - Local-debt Financing'!AY281", "'Input 5 - Local-debt Financing'!AY305", "'Input 5 - Local-debt Financing'!AY471", "'Input 5 - Local-debt Financing'!AY495"), 2028: ("'Input 5 - Local-debt Financing'!AI258", "'Input 5 - Local-debt Financing'!AI282", "'Input 5 - Local-debt Financing'!AI306", "'Input 5 - Local-debt Financing'!AI472", "'Input 5 - Local-debt Financing'!AI496", "'Input 5 - Local-debt Financing'!AK282", "'Input 5 - Local-debt Financing'!AK496", "'Input 5 - Local-debt Financing'!AL282", "'Input 5 - Local-debt Financing'!AL496", "'Input 5 - Local-debt Financing'!AM282", "'Input 5 - Local-debt Financing'!AM496", "'Input 5 - Local-debt Financing'!AN282", "'Input 5 - Local-debt Financing'!AN496", "'Input 5 - Local-debt Financing'!AO282", "'Input 5 - Local-debt Financing'!AO496", "'Input 5 - Local-debt Financing'!AP282", "'Input 5 - Local-debt Financing'!AP496", "'Input 5 - Local-debt Financing'!AQ282", "'Input 5 - Local-debt Financing'!AQ496", "'Input 5 - Local-debt Financing'!AR282", "'Input 5 - Local-debt Financing'!AR496", "'Input 5 - Local-debt Financing'!AS282", "'Input 5 - Local-debt Financing'!AS496", "'Input 5 - Local-debt Financing'!AT282", "'Input 5 - Local-debt Financing'!AT496", "'Input 5 - Local-debt Financing'!AU282", "'Input 5 - Local-debt Financing'!AU496", "'Input 5 - Local-debt Financing'!AV282", "'Input 5 - Local-debt Financing'!AV496", "'Input 5 - Local-debt Financing'!AW282", "'Input 5 - Local-debt Financing'!AW496", "'Input 5 - Local-debt Financing'!AX282", "'Input 5 - Local-debt Financing'!AX496", "'Input 5 - Local-debt Financing'!AY282", "'Input 5 - Local-debt Financing'!AY496"), 2029: ("'Input 5 - Local-debt Financing'!AJ259", "'Input 5 - Local-debt Financing'!AJ283", "'Input 5 - Local-debt Financing'!AJ307", "'Input 5 - Local-debt Financing'!AJ473", "'Input 5 - Local-debt Financing'!AJ497", "'Input 5 - Local-debt Financing'!AL283", "'Input 5 - Local-debt Financing'!AL497", "'Input 5 - Local-debt Financing'!AM283", "'Input 5 - Local-debt Financing'!AM497", "'Input 5 - Local-debt Financing'!AN283", "'Input 5 - Local-debt Financing'!AN497", "'Input 5 - Local-debt Financing'!AO283", "'Input 5 - Local-debt Financing'!AO497", "'Input 5 - Local-debt Financing'!AP283", "'Input 5 - Local-debt Financing'!AP497", "'Input 5 - Local-debt Financing'!AQ283", "'Input 5 - Local-debt Financing'!AQ497", "'Input 5 - Local-debt Financing'!AR283", "'Input 5 - Local-debt Financing'!AR497", "'Input 5 - Local-debt Financing'!AS283", "'Input 5 - Local-debt Financing'!AS497", "'Input 5 - Local-debt Financing'!AT283", "'Input 5 - Local-debt Financing'!AT497", "'Input 5 - Local-debt Financing'!AU283", "'Input 5 - Local-debt Financing'!AU497", "'Input 5 - Local-debt Financing'!AV283", "'Input 5 - Local-debt Financing'!AV497", "'Input 5 - Local-debt Financing'!AW283", "'Input 5 - Local-debt Financing'!AW497", "'Input 5 - Local-debt Financing'!AX283", "'Input 5 - Local-debt Financing'!AX497", "'Input 5 - Local-debt Financing'!AY283", "'Input 5 - Local-debt Financing'!AY497"), 2030: ("'Input 5 - Local-debt Financing'!AK260", "'Input 5 - Local-debt Financing'!AK308", "'Input 5 - Local-debt Financing'!AK474", "'Input 5 - Local-debt Financing'!AM284", "'Input 5 - Local-debt Financing'!AM498", "'Input 5 - Local-debt Financing'!AN284", "'Input 5 - Local-debt Financing'!AN498", "'Input 5 - Local-debt Financing'!AO284", "'Input 5 - Local-debt Financing'!AO498", "'Input 5 - Local-debt Financing'!AP284", "'Input 5 - Local-debt Financing'!AP498", "'Input 5 - Local-debt Financing'!AQ284", "'Input 5 - Local-debt Financing'!AQ498", "'Input 5 - Local-debt Financing'!AR284", "'Input 5 - Local-debt Financing'!AR498", "'Input 5 - Local-debt Financing'!AS284", "'Input 5 - Local-debt Financing'!AS498", "'Input 5 - Local-debt Financing'!AT284", "'Input 5 - Local-debt Financing'!AT498", "'Input 5 - Local-debt Financing'!AU284", "'Input 5 - Local-debt Financing'!AU498", "'Input 5 - Local-debt Financing'!AV284", "'Input 5 - Local-debt Financing'!AV498", "'Input 5 - Local-debt Financing'!AW284", "'Input 5 - Local-debt Financing'!AW498", "'Input 5 - Local-debt Financing'!AX284", "'Input 5 - Local-debt Financing'!AX498", "'Input 5 - Local-debt Financing'!AY284", "'Input 5 - Local-debt Financing'!AY498"), 2031: ("'Input 5 - Local-debt Financing'!AL261", "'Input 5 - Local-debt Financing'!AL309", "'Input 5 - Local-debt Financing'!AL475", "'Input 5 - Local-debt Financing'!AN285", "'Input 5 - Local-debt Financing'!AN499", "'Input 5 - Local-debt Financing'!AO285", "'Input 5 - Local-debt Financing'!AO499", "'Input 5 - Local-debt Financing'!AP285", "'Input 5 - Local-debt Financing'!AP499", "'Input 5 - Local-debt Financing'!AQ285", "'Input 5 - Local-debt Financing'!AQ499", "'Input 5 - Local-debt Financing'!AR285", "'Input 5 - Local-debt Financing'!AR499", "'Input 5 - Local-debt Financing'!AS285", "'Input 5 - Local-debt Financing'!AS499", "'Input 5 - Local-debt Financing'!AT285", "'Input 5 - Local-debt Financing'!AT499", "'Input 5 - Local-debt Financing'!AU285", "'Input 5 - Local-debt Financing'!AU499", "'Input 5 - Local-debt Financing'!AV285", "'Input 5 - Local-debt Financing'!AV499", "'Input 5 - Local-debt Financing'!AW285", "'Input 5 - Local-debt Financing'!AW499", "'Input 5 - Local-debt Financing'!AX285", "'Input 5 - Local-debt Financing'!AX499", "'Input 5 - Local-debt Financing'!AY285", "'Input 5 - Local-debt Financing'!AY499"), 2032: ("'Input 5 - Local-debt Financing'!AM262", "'Input 5 - Local-debt Financing'!AM310", "'Input 5 - Local-debt Financing'!AM476", "'Input 5 - Local-debt Financing'!AO286", "'Input 5 - Local-debt Financing'!AO500", "'Input 5 - Local-debt Financing'!AP286", "'Input 5 - Local-debt Financing'!AP500", "'Input 5 - Local-debt Financing'!AQ286", "'Input 5 - Local-debt Financing'!AQ500", "'Input 5 - Local-debt Financing'!AR286", "'Input 5 - Local-debt Financing'!AR500", "'Input 5 - Local-debt Financing'!AS286", "'Input 5 - Local-debt Financing'!AS500", "'Input 5 - Local-debt Financing'!AT286", "'Input 5 - Local-debt Financing'!AT500", "'Input 5 - Local-debt Financing'!AU286", "'Input 5 - Local-debt Financing'!AU500", "'Input 5 - Local-debt Financing'!AV286", "'Input 5 - Local-debt Financing'!AV500", "'Input 5 - Local-debt Financing'!AW286", "'Input 5 - Local-debt Financing'!AW500", "'Input 5 - Local-debt Financing'!AX286", "'Input 5 - Local-debt Financing'!AX500", "'Input 5 - Local-debt Financing'!AY286", "'Input 5 - Local-debt Financing'!AY500"), 2033: ("'Input 5 - Local-debt Financing'!AN263", "'Input 5 - Local-debt Financing'!AN311", "'Input 5 - Local-debt Financing'!AN477", "'Input 5 - Local-debt Financing'!AP287", "'Input 5 - Local-debt Financing'!AP501", "'Input 5 - Local-debt Financing'!AQ287", "'Input 5 - Local-debt Financing'!AQ501", "'Input 5 - Local-debt Financing'!AR287", "'Input 5 - Local-debt Financing'!AR501", "'Input 5 - Local-debt Financing'!AS287", "'Input 5 - Local-debt Financing'!AS501", "'Input 5 - Local-debt Financing'!AT287", "'Input 5 - Local-debt Financing'!AT501", "'Input 5 - Local-debt Financing'!AU287", "'Input 5 - Local-debt Financing'!AU501", "'Input 5 - Local-debt Financing'!AV287", "'Input 5 - Local-debt Financing'!AV501", "'Input 5 - Local-debt Financing'!AW287", "'Input 5 - Local-debt Financing'!AW501", "'Input 5 - Local-debt Financing'!AX287", "'Input 5 - Local-debt Financing'!AX501", "'Input 5 - Local-debt Financing'!AY287", "'Input 5 - Local-debt Financing'!AY501"), 2034: ("'Input 5 - Local-debt Financing'!AO264", "'Input 5 - Local-debt Financing'!AO312", "'Input 5 - Local-debt Financing'!AO478", "'Input 5 - Local-debt Financing'!AQ288", "'Input 5 - Local-debt Financing'!AQ502", "'Input 5 - Local-debt Financing'!AR288", "'Input 5 - Local-debt Financing'!AR502", "'Input 5 - Local-debt Financing'!AS288", "'Input 5 - Local-debt Financing'!AS502", "'Input 5 - Local-debt Financing'!AT288", "'Input 5 - Local-debt Financing'!AT502", "'Input 5 - Local-debt Financing'!AU288", "'Input 5 - Local-debt Financing'!AU502", "'Input 5 - Local-debt Financing'!AV288", "'Input 5 - Local-debt Financing'!AV502", "'Input 5 - Local-debt Financing'!AW288", "'Input 5 - Local-debt Financing'!AW502", "'Input 5 - Local-debt Financing'!AX288", "'Input 5 - Local-debt Financing'!AX502", "'Input 5 - Local-debt Financing'!AY288", "'Input 5 - Local-debt Financing'!AY502"), 2035: ("'Input 5 - Local-debt Financing'!AP265", "'Input 5 - Local-debt Financing'!AP313", "'Input 5 - Local-debt Financing'!AP479", "'Input 5 - Local-debt Financing'!AR289", "'Input 5 - Local-debt Financing'!AR503", "'Input 5 - Local-debt Financing'!AS289", "'Input 5 - Local-debt Financing'!AS503", "'Input 5 - Local-debt Financing'!AT289", "'Input 5 - Local-debt Financing'!AT503", "'Input 5 - Local-debt Financing'!AU289", "'Input 5 - Local-debt Financing'!AU503", "'Input 5 - Local-debt Financing'!AV289", "'Input 5 - Local-debt Financing'!AV503", "'Input 5 - Local-debt Financing'!AW289", "'Input 5 - Local-debt Financing'!AW503", "'Input 5 - Local-debt Financing'!AX289", "'Input 5 - Local-debt Financing'!AX503", "'Input 5 - Local-debt Financing'!AY289", "'Input 5 - Local-debt Financing'!AY503"), 2036: ("'Input 5 - Local-debt Financing'!AQ266", "'Input 5 - Local-debt Financing'!AQ314", "'Input 5 - Local-debt Financing'!AQ480", "'Input 5 - Local-debt Financing'!AS290", "'Input 5 - Local-debt Financing'!AS504", "'Input 5 - Local-debt Financing'!AT290", "'Input 5 - Local-debt Financing'!AT504", "'Input 5 - Local-debt Financing'!AU290", "'Input 5 - Local-debt Financing'!AU504", "'Input 5 - Local-debt Financing'!AV290", "'Input 5 - Local-debt Financing'!AV504", "'Input 5 - Local-debt Financing'!AW290", "'Input 5 - Local-debt Financing'!AW504", "'Input 5 - Local-debt Financing'!AX290", "'Input 5 - Local-debt Financing'!AX504", "'Input 5 - Local-debt Financing'!AY290", "'Input 5 - Local-debt Financing'!AY504"), 2037: ("'Input 5 - Local-debt Financing'!AR267", "'Input 5 - Local-debt Financing'!AR315", "'Input 5 - Local-debt Financing'!AR481", "'Input 5 - Local-debt Financing'!AT291", "'Input 5 - Local-debt Financing'!AT505", "'Input 5 - Local-debt Financing'!AU291", "'Input 5 - Local-debt Financing'!AU505", "'Input 5 - Local-debt Financing'!AV291", "'Input 5 - Local-debt Financing'!AV505", "'Input 5 - Local-debt Financing'!AW291", "'Input 5 - Local-debt Financing'!AW505", "'Input 5 - Local-debt Financing'!AX291", "'Input 5 - Local-debt Financing'!AX505", "'Input 5 - Local-debt Financing'!AY291", "'Input 5 - Local-debt Financing'!AY505"), 2038: ("'Input 5 - Local-debt Financing'!AS268", "'Input 5 - Local-debt Financing'!AS316", "'Input 5 - Local-debt Financing'!AS482", "'Input 5 - Local-debt Financing'!AU292", "'Input 5 - Local-debt Financing'!AU506", "'Input 5 - Local-debt Financing'!AV292", "'Input 5 - Local-debt Financing'!AV506", "'Input 5 - Local-debt Financing'!AW292", "'Input 5 - Local-debt Financing'!AW506", "'Input 5 - Local-debt Financing'!AX292", "'Input 5 - Local-debt Financing'!AX506", "'Input 5 - Local-debt Financing'!AY292", "'Input 5 - Local-debt Financing'!AY506"), 2039: ("'Input 5 - Local-debt Financing'!AT269", "'Input 5 - Local-debt Financing'!AT317", "'Input 5 - Local-debt Financing'!AT483", "'Input 5 - Local-debt Financing'!AV293", "'Input 5 - Local-debt Financing'!AV507", "'Input 5 - Local-debt Financing'!AW293", "'Input 5 - Local-debt Financing'!AW507", "'Input 5 - Local-debt Financing'!AX293", "'Input 5 - Local-debt Financing'!AX507", "'Input 5 - Local-debt Financing'!AY293", "'Input 5 - Local-debt Financing'!AY507"), 2040: ("'Input 5 - Local-debt Financing'!AU270", "'Input 5 - Local-debt Financing'!AU318", "'Input 5 - Local-debt Financing'!AU484", "'Input 5 - Local-debt Financing'!AW294", "'Input 5 - Local-debt Financing'!AW508", "'Input 5 - Local-debt Financing'!AX294", "'Input 5 - Local-debt Financing'!AX508", "'Input 5 - Local-debt Financing'!AY294", "'Input 5 - Local-debt Financing'!AY508"), 2041: ("'Input 5 - Local-debt Financing'!AV271", "'Input 5 - Local-debt Financing'!AV319", "'Input 5 - Local-debt Financing'!AV485", "'Input 5 - Local-debt Financing'!AX295", "'Input 5 - Local-debt Financing'!AX509", "'Input 5 - Local-debt Financing'!AY295", "'Input 5 - Local-debt Financing'!AY509"), 2042: ("'Input 5 - Local-debt Financing'!AW272", "'Input 5 - Local-debt Financing'!AW320", "'Input 5 - Local-debt Financing'!AW486", "'Input 5 - Local-debt Financing'!AY296", "'Input 5 - Local-debt Financing'!AY510"), 2043: ("'Input 5 - Local-debt Financing'!AX273", "'Input 5 - Local-debt Financing'!AX321", "'Input 5 - Local-debt Financing'!AX487")},
            values=values, start_year=start_year, strict=strict,
        )
