#!/usr/bin/env python3
"""Reshape-to-keys prototype: lift Input 4 series onto INSTRUMENT (and TIME_PERIOD) keys.

Reads original shards under /workspace/bindings/ (not modified) and writes:

- bindings/remodeled/input4.bindings.yaml
- bindings/remodeled/input4-audit.md
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import yaml
from openpyxl import load_workbook

from input4_cells import (
    bbox_a1,
    cell_addr,
    cells_of_range,
    col_to_num,
    compact_row_specs,
    expand_a1,
    num_to_col,
    split_sheet_range,
)

ROOT = Path("/workspace")
BINDINGS = ROOT / "bindings"
OUT_DIR = BINDINGS / "remodeled"
SOURCES = [
    BINDINGS / "inputs.bindings.yaml",
    BINDINGS / "constants.bindings.yaml",
    BINDINGS / "internals-graph-coverage.bindings.yaml",
    BINDINGS / "internals-rest.bindings.yaml",
]
SHEET = "Input 4 - External Financing"
WORKBOOK = ROOT / "data" / "lic-dsf-template-2025-08-12.xlsx"

# PV_Base / terms instrument rows on Input 4 (loaders skip incomplete PC2–PC5).
TERM_ROWS = {
    10, 11, 12, 13, 14, 15, 16, 17, 18, 19,
    21, 22, 23,
    26, 27, 28, 29, 30,
    32, 33, 34, 35, 36,
    38, 39, 40, 41, 42,
    45,
    49, 50, 51,
    54, 55, 56,
    59, 60, 61,
}
IDA_SCALE_ROWS = set(range(67, 76))
DISBURSE_COLS = set(range(col_to_num("L"), col_to_num("AF") + 1))
TERMS_FH_COLS = {col_to_num("F"), col_to_num("G"), col_to_num("H")}
DISCOUNT_COL = col_to_num("E")
FX_NR_ROWS = {54, 55, 56}
FX_R_ROWS = {59, 60, 61}

# Instrument labels that appear as draft-pass INDICATOR / PARAMETER values.
KNOWN_INSTRUMENTS = {
    "IMF",
    "IDA - regular",
    "IDA - 50Y loans",
    "IDA - SML",
    "IDA NEW 40-year credits",
    "IDA NEW Regular",
    "IDA NEW Blend (also enter) -->",
    "IDA NEW 60-year credits",
    "MULTI1",
    "MULTI2",
    "OTH_MULTI1",
    "OTH_MULTI2",
    "OTH_MULTI3",
    "Export Credit Agencies",
    "PC2",
    "PC3",
    "PC4",
    "PC5",
    "Export Import Bank of NPC",
    "NPC2",
    "NPC3",
    "NPC4",
    "NPC5",
    "Eurobond",
    "Commecial Bank",
    "COM3",
    "COM4",
    "COM5",
    "PPG ST external debt",
    "Bonds (1 to 3 years)-LC",
    "Bonds (4 to 7 years)-LC",
    "Bonds (beyond 7 years)-LC",
    "Bonds (1 to 3 years)-FX",
    "Bonds (4 to 7 years)-FX",
    "Bonds (beyond 7 years)-FX",
    "IDA - small economy",
    "IDA - blend",
}


def direction_of(series: dict) -> str:
    for d in ("input", "internal", "constant", "output"):
        if d in series:
            return d
    return "none"


def load_originals() -> tuple[dict, list[dict]]:
    concept_scheme = None
    rows: list[dict] = []
    for path in SOURCES:
        doc = yaml.safe_load(path.read_text())
        if concept_scheme is None and path.name == "inputs.bindings.yaml":
            concept_scheme = doc["concept_scheme"]
        for s in doc.get("series") or []:
            sheet = s.get("sheet") or ""
            dr = s.get("data_range") or ""
            sheet2, a1 = split_sheet_range(dr)
            if not (str(sheet).startswith("Input 4") or str(sheet2 or "").startswith("Input 4")):
                continue
            cells = expand_a1(a1)
            ctx = s.get("series_context") or {}
            struct = s.get("structure") or {}
            measure = struct.get("measure") or {}
            dims = struct.get("dimensions") or []
            tp = next((d for d in dims if d.get("concept") == "TIME_PERIOD"), None)
            tp_bind = (tp or {}).get("bind") or {}
            rec = {
                "source": path.name,
                "raw": s,
                "id": s["id"],
                "sheet": sheet,
                "data_range": dr,
                "a1": a1,
                "layout": s.get("layout"),
                "direction": direction_of(s),
                "ctx": ctx,
                "INDICATOR": ctx.get("INDICATOR"),
                "TABLE": ctx.get("TABLE"),
                "VARIANT": ctx.get("VARIANT"),
                "INSTRUMENT": ctx.get("INSTRUMENT"),
                "HOLDER": ctx.get("HOLDER"),
                "PARAMETER": ctx.get("PARAMETER"),
                "dtype": measure.get("dtype"),
                "measure_read": (measure.get("bind") or {}).get("read"),
                "dims": dims,
                "key": list(s.get("key") or []),
                "tp_header_row": tp_bind.get("header_row"),
                "tp_kind": tp_bind.get("kind"),
                "cells": cells,
                "rows": sorted({r for _, r in cells}),
                "cols": sorted({c for c, _ in cells}),
                "n_cells": len(cells),
                "input": s.get("input"),
                "notes": s.get("notes"),
            }
            rows.append(rec)
    if concept_scheme is None:
        raise SystemExit("concept_scheme missing from inputs.bindings.yaml")
    return concept_scheme, rows


def load_col_b(rows_needed: set[int]) -> dict[int, str]:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb[SHEET]
    out: dict[int, str] = {}
    for row in sorted(rows_needed):
        val = ws.cell(row, 2).value
        if val is None:
            continue
        text = str(val).strip()
        if text:
            out[row] = text
    wb.close()
    return out


def instrument_of(rec: dict, col_b: dict[int, str]) -> str | None:
    if rec["INSTRUMENT"]:
        return rec["INSTRUMENT"]
    ind = rec["INDICATOR"]
    if ind in KNOWN_INSTRUMENTS:
        return ind
    par = rec["PARAMETER"]
    if par in KNOWN_INSTRUMENTS:
        return par
    # Single-row leftover: prefer column B.
    if len(rec["rows"]) == 1:
        label = col_b.get(rec["rows"][0])
        if label in KNOWN_INSTRUMENTS:
            return label
    return None


def is_terms_fh(rec: dict) -> bool:
    return bool(rec["cols"]) and set(rec["cols"]) <= TERMS_FH_COLS and rec["direction"] == "internal"


def is_discount_e(rec: dict) -> bool:
    return rec["cols"] == [DISCOUNT_COL] and rec["direction"] == "internal"


def is_disbursement_geometry(rec: dict) -> bool:
    if rec["direction"] not in {"internal", "input"}:
        return False
    if not rec["cols"]:
        return False
    if not set(rec["cols"]) <= DISBURSE_COLS:
        return False
    if not set(rec["rows"]) <= TERM_ROWS:
        return False
    return True


def is_ida_scale_principal_geometry(rec: dict) -> bool:
    if rec["direction"] not in {"internal", "constant"}:
        return False
    if not set(rec["rows"]) <= IDA_SCALE_ROWS:
        return False
    # Principal schedule lives to the right of terms columns on the IDA block.
    if min(rec["cols"]) < col_to_num("H"):
        return False
    ind = rec["INDICATOR"]
    if ind == "Principal repayment":
        return True
    if ind in KNOWN_INSTRUMENTS or rec["PARAMETER"] in KNOWN_INSTRUMENTS:
        return True
    if rec["id"].startswith("in4_extfin_ida_s_lending_terms") or rec["id"].startswith("in4_extfin_ida_blend"):
        return True
    return False


def classify(rec: dict, col_b: dict[int, str]) -> str:
    """Return merge-group id (or unlifted_* / one-off id)."""
    rid = rec["id"]
    ind = rec["INDICATOR"]
    table = rec["TABLE"]
    direction = rec["direction"]
    variant = rec["VARIANT"]
    a1 = rec["a1"]

    # Explicit unique one-offs.
    if a1 == "D16" or ind == "blend_variant":
        return "input4_blend_variant"
    if a1 == "G6":
        return "input4_grace_period_column_header"
    if a1 == "C74" or ind == "fixed_label":
        return "input4_ida_scale_blend_fixed"
    if a1 == "D11" or ind == "translated_name":
        return "input4_ida_regular_translated_name"
    if a1 == "D74" and variant == "fixed":
        return "input4_ida_scale_blend_fixed_interest"
    if rec["rows"] == [8] and set(rec["cols"]) <= DISBURSE_COLS:
        return "input4_disbursement_totals"

    if direction == "input":
        if ind == "Interest rate":
            return "input4_interest_rate"
        if ind == "Grace period":
            return "input4_grace_period"
        if ind == "Loan maturity":
            return "input4_loan_maturity"
        if table == "input_4.disbursements" or (
            rid.startswith("input4_disbursements_") and rec["tp_header_row"] == 6
        ):
            return "input4_disbursements"
        if ind == "blend_scale_key":
            return "input4_blend_scale_key"
        if ind == "instrument_name":
            return "input4_instrument_names"

    if direction == "constant":
        if table == "input_4.ida_scale" or str(table).startswith("input_4.ida"):
            if ind == "scale_name":
                return "input4_ida_scale_name"
            if ind == "Service fee and interest":
                return "input4_ida_scale_service_fee"
            if ind == "Grace period":
                return "input4_ida_scale_grace"
            if ind in {"Loan Maturity", "Loan maturity"}:
                return "input4_ida_scale_maturity"
            if ind == "Principal repayment":
                return "input4_ida_scale_principal"

    # Internal F:H terms (Interest / Grace / Maturity column headers on row 6).
    if is_terms_fh(rec):
        rows = set(rec["rows"])
        if rows <= FX_NR_ROWS:
            return "input4_terms_bonds_fx_non_residents"
        if rows <= FX_R_ROWS:
            return "input4_terms_bonds_fx_residents"
        if table == "input_4.terms":
            return "input4_terms_ida_regular"
        return "input4_terms_ida_and_lc"

    # Column E discount rate on terms rows.
    if is_discount_e(rec):
        rows = set(rec["rows"])
        if rows <= FX_NR_ROWS:
            return "input4_discount_rate_bonds_fx_non_residents"
        if rows <= FX_R_ROWS:
            return "input4_discount_rate_bonds_fx_residents"
        if rows <= TERM_ROWS:
            return "input4_discount_rate"

    # Input-vs-internal disbursements on the year grid (header row 6).
    if direction == "internal" and is_disbursement_geometry(rec):
        return "input4_disbursements_internal"

    # IDA scale principal leftovers (graph coverage + authored internal).
    if direction == "internal" and is_ida_scale_principal_geometry(rec):
        return "input4_ida_scale_principal_internal"

    return f"input4_unlifted_{rid}"


def instrument_rows_for(rec: dict, col_b: dict[int, str]) -> dict[str, list[int]]:
    """Map instrument label → member rows, expanding row_label blocks via column B.

    Draft leftovers often store only the first row's label as INDICATOR while the
    range covers several instruments (E10:E19, E54:E56, …). When column B has
    distinct known instrument names, those labels win.
    """
    rows = list(rec["rows"])
    b_labels = [col_b.get(r) for r in rows]
    distinct_instruments = [lab for lab in b_labels if lab in KNOWN_INSTRUMENTS]
    if len(rows) > 1 and len(set(distinct_instruments)) > 1:
        mapping: dict[str, list[int]] = defaultdict(list)
        for row, label in zip(rows, b_labels):
            if label in KNOWN_INSTRUMENTS:
                mapping[label].append(row)
            elif rec["INSTRUMENT"]:
                mapping[rec["INSTRUMENT"]].append(row)
            else:
                mapping[label or f"row_{row}"].append(row)
        return dict(mapping)
    inst = rec["INSTRUMENT"]
    if inst:
        return {inst: list(rows)}
    if rec["INDICATOR"] in KNOWN_INSTRUMENTS:
        return {rec["INDICATOR"]: list(rows)}
    if rec["PARAMETER"] in KNOWN_INSTRUMENTS:
        return {rec["PARAMETER"]: list(rows)}
    mapping = defaultdict(list)
    for row in rows:
        label = col_b.get(row)
        if label and label in KNOWN_INSTRUMENTS:
            mapping[label].append(row)
        elif label:
            mapping[label].append(row)
        else:
            mapping[f"row_{row}"].append(row)
    return dict(mapping)


def merge_value_map(members: list[dict], col_b: dict[int, str]) -> dict[str, list[int]]:
    merged: dict[str, list[int]] = defaultdict(list)
    for rec in members:
        for inst, rows in instrument_rows_for(rec, col_b).items():
            merged[inst].extend(rows)
    for inst in list(merged):
        merged[inst] = sorted(set(merged[inst]))
    return dict(merged)


def value_map_values(mapping: dict[str, list[int]]) -> dict[str, int | str]:
    out: dict[str, int | str] = {}
    for inst, rows in mapping.items():
        specs = compact_row_specs(rows)
        if len(specs) != 1:
            # Non-contiguous: keep the compact list joined — schema wants one spec.
            # Caller should have split such instruments (e.g. FX bonds by HOLDER).
            raise ValueError(f"non-contiguous rows for {inst!r}: {specs}")
        spec = specs[0]
        out[inst] = spec
    return out


def exclude_for(cells: list[tuple[int, int]], member_rows: set[int]) -> list[int | str]:
    if not cells:
        return []
    _, r1, _, r2 = (
        min(c for c, _ in cells),
        min(r for _, r in cells),
        max(c for c, _ in cells),
        max(r for _, r in cells),
    )
    missing = [r for r in range(r1, r2 + 1) if r not in member_rows]
    return compact_row_specs(missing)


def domains_of(members: list[dict]) -> dict | None:
    doms = []
    for rec in members:
        inp = rec.get("input") or {}
        doms.append(inp.get("domain"))
    first = json.dumps(doms[0], sort_keys=True)
    if all(json.dumps(d, sort_keys=True) == first for d in doms):
        return doms[0]
    return "__DIVERGENT__"


def sheet_range(a1: str) -> str:
    return f"'{SHEET}'!{a1}"


def yaml_scalar(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return str(value)
    text = str(value)
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", text):
        return text
    return json.dumps(text, ensure_ascii=False)


def emit_value_map(values: dict[str, int | str], indent: str) -> list[str]:
    lines = [f"{indent}values:"]
    for key, spec in values.items():
        lines.append(f"{indent}  {yaml_scalar(key)}: {yaml_scalar(spec)}")
    return lines


def emit_series(s: dict) -> str:
    lines: list[str] = [f"- id: {s['id']}"]
    lines.append(f"  sheet: {yaml_scalar(SHEET)}")
    lines.append(f"  data_range: {yaml_scalar(s['data_range'])}")
    lines.append(f"  layout: {s['layout']}")
    direction = s["direction"]
    if direction == "input":
        lines.append("  input:")
        lines.append("    setter:")
        lines.append(f"      name: {s['setter']}")
        lines.append("      record_contract: records")
        lines.append("      strict: true")
        if s.get("domain"):
            dom = s["domain"]
            if "between" in dom:
                lines.append("    domain:")
                lines.append("      between:")
                lines.append(f"        min: {dom['between']['min']}")
                lines.append(f"        max: {dom['between']['max']}")
            elif "real_between" in dom:
                lines.append("    domain:")
                lines.append("      real_between:")
                lines.append(f"        min: {dom['real_between']['min']}")
                lines.append(f"        max: {dom['real_between']['max']}")
            elif "enum" in dom:
                lines.append("    domain:")
                lines.append("      enum:")
                for item in dom["enum"]:
                    lines.append(f"      - {yaml_scalar(item)}")
    elif direction == "internal":
        lines.append("  internal: {}")
    elif direction == "constant":
        lines.append("  constant: {}")
    if s.get("exclude_rows"):
        lines.append("  exclude_rows:")
        for spec in s["exclude_rows"]:
            lines.append(f"  - {yaml_scalar(spec)}")
    lines.append("  structure:")
    lines.append("    measure:")
    lines.append("      concept: OBS_VALUE")
    lines.append(f"      dtype: {s['dtype']}")
    lines.append("      bind:")
    lines.append("        kind: data_cell")
    lines.append(f"        read: {s['read']}")
    if not s["dimensions"]:
        lines.append("    dimensions: []")
    else:
        lines.append("    dimensions:")
        for dim in s["dimensions"]:
            lines.append(f"    - id: {dim['id']}")
            lines.append(f"      concept: {dim['concept']}")
            lines.append(f"      role: {dim['role']}")
            lines.append(f"      scope: {dim['scope']}")
            bind = dim["bind"]
            lines.append("      bind:")
            lines.append(f"        kind: {bind['kind']}")
            if bind["kind"] == "value_map":
                lines.extend(emit_value_map(bind["values"], "        "))
            elif bind["kind"] == "column_header":
                lines.append(f"        header_row: {bind['header_row']}")
                lines.append(f"        read: {bind['read']}")
            elif bind["kind"] == "constant":
                lines.append(f"        value: {yaml_scalar(bind['value'])}")
            elif bind["kind"] == "row_label":
                lines.append(f"        label_column: {bind['label_column']}")
                lines.append(f"        read: {bind['read']}")
                if bind.get("normalize"):
                    lines.append(f"        normalize: {bind['normalize']}")
            elif bind["kind"] == "data_cell":
                lines.append(f"        read: {bind['read']}")
            elif bind["kind"] == "cell":
                if "address" in bind:
                    lines.append(f"        address: {yaml_scalar(bind['address'])}")
    if not s["key"]:
        lines.append("  key: []")
    else:
        lines.append("  key:")
        for k in s["key"]:
            lines.append(f"  - {k}")
    ctx = s.get("series_context") or {}
    if ctx:
        lines.append("  series_context:")
        for ck, cv in ctx.items():
            lines.append(f"    {ck}: {yaml_scalar(cv)}")
    if s.get("notes"):
        lines.append(f"  notes: {yaml_scalar(s['notes'])}")
    return "\n".join(lines)


def emit_concept_scheme(scheme: dict) -> str:
    lines = ["concept_scheme:", f"  id: {scheme['id']}", "  concepts:"]
    for c in scheme["concepts"]:
        lines.append(f"  - id: {c['id']}")
        lines.append(f"    name: {c['name']}")
        lines.append(f"    dtype: {c['dtype']}")
        if c.get("sdmx_concept"):
            lines.append(f"    sdmx_concept: {c['sdmx_concept']}")
    return "\n".join(lines)


def instrument_dim(values: dict[str, int | str]) -> dict:
    return {
        "id": "INSTRUMENT",
        "concept": "INSTRUMENT",
        "role": "key",
        "scope": "cell",
        "bind": {"kind": "value_map", "values": values},
    }


def time_dim(header_row: int) -> dict:
    return {
        "id": "TIME_PERIOD",
        "concept": "TIME_PERIOD",
        "role": "key",
        "scope": "cell",
        "bind": {"kind": "column_header", "header_row": header_row, "read": "int"},
    }


def indicator_col_header(header_row: int) -> dict:
    return {
        "id": "INDICATOR",
        "concept": "INDICATOR",
        "role": "key",
        "scope": "cell",
        "bind": {"kind": "column_header", "header_row": header_row, "read": "string"},
    }


def holder_constant(value: str) -> dict:
    return {
        "id": "HOLDER",
        "concept": "HOLDER",
        "role": "key",
        "scope": "cell",
        "bind": {"kind": "constant", "value": value},
    }


NOTES = {
    "input4_interest_rate": "Interest rates on new external PPG borrowing, keyed by instrument.",
    "input4_grace_period": "Grace periods in years on new external PPG borrowing, keyed by instrument.",
    "input4_loan_maturity": "Loan maturities in years on new external PPG borrowing, keyed by instrument.",
    "input4_disbursements": "User-entered new-loan disbursements by instrument and projection year.",
    "input4_blend_scale_key": "IDA scale lookup keys for instruments that copy published IDA terms.",
    "input4_instrument_names": "User-editable IDA instrument names on the external financing list.",
    "input4_blend_variant": "Whether IDA NEW Blend uses floating or fixed terms.",
    "input4_ida_scale_name": "Published IDA scale names on the lending-terms block.",
    "input4_ida_scale_service_fee": "Published IDA service fee and interest, keyed by instrument.",
    "input4_ida_scale_grace": "Published IDA grace periods, keyed by instrument.",
    "input4_ida_scale_maturity": "Published IDA loan maturities, keyed by instrument.",
    "input4_ida_scale_principal": "Published IDA principal-repayment shares by year of loan life.",
    "input4_ida_scale_blend_fixed": "Fixed-blend label on the IDA NEW Blend scale row.",
    "input4_disbursements_internal": "Formula disbursements by instrument and projection year (non-input cells).",
    "input4_discount_rate": "Discount rate applied to each external instrument row.",
    "input4_discount_rate_bonds_fx_non_residents": "Discount rate on FX local bonds held by non-residents.",
    "input4_discount_rate_bonds_fx_residents": "Discount rate on FX local bonds held by residents.",
    "input4_terms_ida_regular": "Computed interest, grace, and maturity for IDA regular.",
    "input4_terms_ida_and_lc": "Computed interest, grace, and maturity for IDA windows and LC bonds.",
    "input4_terms_bonds_fx_non_residents": "Computed interest, grace, and maturity for FX local bonds held by non-residents.",
    "input4_terms_bonds_fx_residents": "Computed interest, grace, and maturity for FX local bonds held by residents.",
    "input4_ida_scale_principal_internal": "Formula IDA principal-repayment shares by year of loan life.",
    "input4_disbursement_totals": "Row-total new external disbursements by projection year.",
    "input4_grace_period_column_header": "Column header text for grace period on the terms grid.",
    "input4_ida_regular_translated_name": "Translated display name for IDA regular.",
    "input4_ida_scale_blend_fixed_interest": "Fixed-path interest placeholder on the IDA NEW Blend scale row.",
}

# Preferred emission order for named groups (unlifted follow at the end).
GROUP_ORDER = [
    "input4_interest_rate",
    "input4_grace_period",
    "input4_loan_maturity",
    "input4_disbursements",
    "input4_instrument_names",
    "input4_blend_variant",
    "input4_blend_scale_key",
    "input4_ida_scale_name",
    "input4_ida_scale_service_fee",
    "input4_ida_scale_grace",
    "input4_ida_scale_maturity",
    "input4_ida_scale_principal",
    "input4_ida_scale_blend_fixed",
    "input4_discount_rate",
    "input4_discount_rate_bonds_fx_non_residents",
    "input4_discount_rate_bonds_fx_residents",
    "input4_terms_ida_regular",
    "input4_terms_ida_and_lc",
    "input4_terms_bonds_fx_non_residents",
    "input4_terms_bonds_fx_residents",
    "input4_disbursements_internal",
    "input4_ida_scale_principal_internal",
    "input4_disbursement_totals",
    "input4_grace_period_column_header",
    "input4_ida_regular_translated_name",
    "input4_ida_scale_blend_fixed_interest",
]


def dtype_and_read(members: list[dict], *, unify_float: bool = False) -> tuple[str, str]:
    dtypes = {m["dtype"] for m in members}
    reads = {m["measure_read"] or m["dtype"] for m in members}
    if unify_float and dtypes <= {"int", "float"}:
        return "float", "float"
    if len(dtypes) != 1:
        # Prefer float if mixed numeric, else first.
        if dtypes <= {"int", "float"}:
            return "float", "float"
        return next(iter(dtypes)), next(iter(reads))
    dtype = next(iter(dtypes))
    read = next(iter(reads))
    return dtype, read


def build_group_series(
    gid: str,
    members: list[dict],
    col_b: dict[int, str],
    audit: dict,
) -> dict:
    all_cells = [cell for m in members for cell in m["cells"]]
    member_rows = {row for m in members for row in m["rows"]}
    a1 = bbox_a1(all_cells)
    exclude = exclude_for(all_cells, member_rows)
    direction = members[0]["direction"]
    tables = {m["TABLE"] for m in members if m["TABLE"]}
    indicators = {m["INDICATOR"] for m in members if m["INDICATOR"] and m["INDICATOR"] not in KNOWN_INSTRUMENTS}
    holders = {m["HOLDER"] for m in members if m["HOLDER"]}
    tp_headers = {m["tp_header_row"] for m in members if m["tp_header_row"] is not None}

    has_time = any(m["tp_kind"] == "column_header" or m["tp_header_row"] is not None for m in members)
    # Absorb scalar L-column disbursement leftovers onto the year grid.
    if gid == "input4_disbursements_internal":
        has_time = True
        tp_headers = {6}
    if gid == "input4_ida_scale_principal_internal":
        has_time = True
        original_headers = sorted(tp_headers) if tp_headers else []
        member_header_rows = sorted({(m["id"], m["tp_header_row"]) for m in members})
        if original_headers != [66]:
            audit.setdefault("header_row_corrections", []).append(
                {
                    "id": gid,
                    "original": original_headers or ["(none on scalars)"],
                    "remodeled": 66,
                    "reason": "IDA scale principal uses year-of-loan-life headers on row 66; graph-coverage leftovers were bound to disbursement years on row 6.",
                    "member_headers": member_header_rows,
                }
            )
        tp_headers = {66}

    if len(tp_headers) > 1:
        raise ValueError(f"{gid}: mixed TIME_PERIOD header_row {tp_headers}")

    unify_float = gid in {
        "input4_disbursements_internal",
        "input4_ida_scale_principal_internal",
        "input4_terms_ida_and_lc",
        "input4_terms_ida_regular",
        "input4_terms_bonds_fx_non_residents",
        "input4_terms_bonds_fx_residents",
    }
    dtype, read = dtype_and_read(members, unify_float=unify_float)
    orig_dtypes = sorted({m["dtype"] for m in members})
    if len(orig_dtypes) > 1:
        audit.setdefault("dtype_unifications", []).append(
            {"id": gid, "original": orig_dtypes, "remodeled": dtype}
        )

    mapping = merge_value_map(members, col_b)
    # Groups that are already one block of names (instrument_names, scale_name, grace/maturity)
    # still want an INSTRUMENT value_map from member rows / column B.
    if gid in {"input4_instrument_names", "input4_ida_scale_name", "input4_ida_scale_grace", "input4_ida_scale_maturity"}:
        mapping = {}
        for rec in members:
            for row in rec["rows"]:
                label = rec["INSTRUMENT"] or col_b.get(row)
                if not label:
                    raise ValueError(f"{gid}: no instrument label for row {row}")
                mapping.setdefault(label, []).append(row)
        mapping = {k: sorted(set(v)) for k, v in mapping.items()}

    vm = value_map_values(mapping) if mapping else {}
    member_rows_from_map = {r for rows in mapping.values() for r in rows}
    if member_rows_from_map:
        exclude = exclude_for(all_cells, member_rows_from_map)

    uses_indicator_header = gid.startswith("input4_terms_")
    one_offs = {
        "input4_blend_variant",
        "input4_ida_scale_blend_fixed",
        "input4_grace_period_column_header",
        "input4_ida_regular_translated_name",
        "input4_ida_scale_blend_fixed_interest",
        "input4_disbursement_totals",
    }

    dims: list[dict] = []
    key: list[str] = []
    if vm and gid not in one_offs:
        dims.append(instrument_dim(vm))
        key.append("INSTRUMENT")

    if gid == "input4_terms_bonds_fx_non_residents" or gid == "input4_discount_rate_bonds_fx_non_residents":
        dims.append(holder_constant("non-residents"))
        key.append("HOLDER")
    if gid == "input4_terms_bonds_fx_residents" or gid == "input4_discount_rate_bonds_fx_residents":
        dims.append(holder_constant("residents"))
        key.append("HOLDER")

    if uses_indicator_header:
        dims.append(indicator_col_header(6))
        key.append("INDICATOR")

    if has_time and gid not in {"input4_grace_period_column_header"}:
        header_row = next(iter(tp_headers)) if tp_headers else (66 if "principal" in gid else 6)
        dims.append(time_dim(header_row))
        key.append("TIME_PERIOD")

    if gid in {
        "input4_blend_variant",
        "input4_ida_scale_blend_fixed",
        "input4_grace_period_column_header",
        "input4_ida_regular_translated_name",
        "input4_ida_scale_blend_fixed_interest",
    }:
        layout = "scalar"
        dims = []
        key = []
    elif len(key) >= 2:
        layout = "matrix"
    else:
        layout = "series"

    # Context
    ctx: dict[str, str] = {}
    if gid.startswith("input4_ida_scale"):
        ctx["TABLE"] = "input_4.ida_scale"
    elif gid in {"input4_disbursements", "input4_disbursements_internal", "input4_disbursement_totals"}:
        ctx["TABLE"] = "input_4.disbursements"
    elif gid.startswith("input4_discount_rate") and gid != "input4_discount_rate" and "bonds" in gid:
        ctx["TABLE"] = "input_4.terms"
    elif gid == "input4_discount_rate":
        ctx["TABLE"] = "input_4.terms"
    elif gid.startswith("input4_terms_ida_and_lc"):
        # Original TABLE was input_4_external_financing.discount_rate (draft mislabel of F:H terms).
        ctx["TABLE"] = "input_4.terms"
    elif gid.startswith("input4_terms_"):
        ctx["TABLE"] = "input_4.terms"
    elif tables:
        ctx["TABLE"] = next(iter(tables)) if len(tables) == 1 else next(iter(sorted(t for t in tables if t)))

    indicator_ctx = {
        "input4_interest_rate": "Interest rate",
        "input4_grace_period": "Grace period",
        "input4_loan_maturity": "Loan maturity",
        "input4_disbursements": "Disbursements",
        "input4_disbursements_internal": "Disbursements",
        "input4_disbursement_totals": "Total disbursements",
        "input4_blend_scale_key": "blend_scale_key",
        "input4_instrument_names": "instrument_name",
        "input4_blend_variant": "blend_variant",
        "input4_ida_scale_name": "scale_name",
        "input4_ida_scale_service_fee": "Service fee and interest",
        "input4_ida_scale_grace": "Grace period",
        "input4_ida_scale_maturity": "Loan Maturity",
        "input4_ida_scale_principal": "Principal repayment",
        "input4_ida_scale_principal_internal": "Principal repayment",
        "input4_ida_scale_blend_fixed": "fixed_label",
        "input4_discount_rate": "Discount rate",
        "input4_discount_rate_bonds_fx_non_residents": "Discount rate",
        "input4_discount_rate_bonds_fx_residents": "Discount rate",
        "input4_grace_period_column_header": "Grace period",
        "input4_ida_regular_translated_name": "translated_name",
        "input4_ida_scale_blend_fixed_interest": "Interest rate",
    }
    if gid in indicator_ctx:
        ctx["INDICATOR"] = indicator_ctx[gid]
    elif len(indicators) == 1:
        ctx["INDICATOR"] = next(iter(indicators))

    if gid == "input4_grace_period_column_header":
        ctx["VARIANT"] = "column_header"
    if gid == "input4_ida_scale_blend_fixed_interest":
        ctx["VARIANT"] = "fixed"
        ctx["INSTRUMENT"] = "IDA NEW Blend (also enter) -->"
    if gid == "input4_blend_variant":
        ctx["INSTRUMENT"] = "IDA NEW Blend (also enter) -->"
    if gid == "input4_ida_scale_blend_fixed":
        ctx["INSTRUMENT"] = "IDA NEW Blend (also enter) -->"
    if gid == "input4_ida_regular_translated_name":
        ctx["INSTRUMENT"] = "IDA - regular"

    series = {
        "id": gid,
        "data_range": sheet_range(a1),
        "layout": layout,
        "direction": direction,
        "dtype": dtype,
        "read": read,
        "dimensions": dims,
        "key": key,
        "series_context": ctx,
        "notes": NOTES.get(gid, f"Remodeled Input 4 series {gid.replace('_', ' ')}."),
        "exclude_rows": exclude,
        "members": [m["id"] for m in members],
        "n_original": len(members),
        "n_original_cells": sum(m["n_cells"] for m in members),
    }
    if direction == "input":
        series["setter"] = f"set_{gid}"
        dom = domains_of(members)
        if dom == "__DIVERGENT__":
            audit.setdefault("divergent_domains", []).append(gid)
        elif dom:
            series["domain"] = dom
    return series


def build_unlifted(rec: dict) -> dict:
    raw = rec["raw"]
    gid = f"input4_unlifted_{rec['id']}"
    direction = rec["direction"]
    struct = raw.get("structure") or {}
    measure = struct.get("measure") or {}
    dims = []
    for d in struct.get("dimensions") or []:
        bind = dict(d.get("bind") or {})
        dims.append(
            {
                "id": d.get("id") or d.get("concept"),
                "concept": d.get("concept"),
                "role": d.get("role") or "key",
                "scope": d.get("scope") or "cell",
                "bind": bind,
            }
        )
    ctx = dict(rec["ctx"])
    series = {
        "id": gid,
        "data_range": rec["data_range"] if rec["data_range"].startswith("'") or "!" in rec["data_range"] else sheet_range(rec["a1"]),
        "layout": rec["layout"] or "scalar",
        "direction": direction,
        "dtype": rec["dtype"] or "float",
        "read": rec["measure_read"] or rec["dtype"] or "float",
        "dimensions": dims,
        "key": list(raw.get("key") or []),
        "series_context": ctx,
        "notes": rec["notes"] or f"Unlifted original series {rec['id']}.",
        "exclude_rows": raw.get("exclude_rows") or [],
        "members": [rec["id"]],
        "n_original": 1,
        "n_original_cells": rec["n_cells"],
    }
    if direction == "input":
        setter = ((raw.get("input") or {}).get("setter") or {}).get("name")
        series["setter"] = setter or f"set_{gid}"
        dom = (raw.get("input") or {}).get("domain")
        if dom:
            series["domain"] = dom
    # Rewrite notes if they are A1-identity; keep semantic original notes.
    if series["notes"] and re.search(r"![A-Z]+\d+", series["notes"]):
        series["notes"] = f"Unlifted original series {rec['id']}."
    return series


def coverage(originals: list[dict], remodeled: list[dict]) -> dict:
    orig_cells: set[str] = set()
    for rec in originals:
        orig_cells |= {cell_addr(c, r) for c, r in rec["cells"]}
    rem_cells: set[str] = set()
    extra_by_series: list[tuple[str, int]] = []
    per_series = []
    for s in remodeled:
        cells = cells_of_range(s["data_range"], s.get("exclude_rows"))
        rem_cells |= cells
        extra_n = len(cells - orig_cells)
        extra_by_series.append((s["id"], extra_n, len(cells)))
        per_series.append((s["id"], len(cells)))
    missing = sorted(orig_cells - rem_cells)
    extra = sorted(rem_cells - orig_cells)
    return {
        "original_cells": len(orig_cells),
        "remodeled_cells": len(rem_cells),
        "missing": missing,
        "extra": extra,
        "per_series": per_series,
        "extra_by_series": extra_by_series,
    }


def write_audit(
    originals: list[dict],
    groups: dict[str, list[dict]],
    remodeled: list[dict],
    cov: dict,
    audit: dict,
) -> str:
    by_src = defaultdict(int)
    by_dir = defaultdict(int)
    for rec in originals:
        by_src[rec["source"]] += 1
        by_dir[rec["direction"]] += 1

    unlifted = [s for s in remodeled if s["id"].startswith("input4_unlifted_")]
    named = [s for s in remodeled if not s["id"].startswith("input4_unlifted_")]

    lines = [
        "# Input 4 reshape-to-keys audit",
        "",
        "Prototype catalog: `bindings/remodeled/input4.bindings.yaml` (schema 1.13.0).",
        "Original shards were not modified.",
        "",
        "## Counts",
        "",
        f"- Original series: **{len(originals)}**",
        f"  - inputs.bindings.yaml: {by_src['inputs.bindings.yaml']}",
        f"  - constants.bindings.yaml: {by_src['constants.bindings.yaml']}",
        f"  - internals-graph-coverage.bindings.yaml: {by_src['internals-graph-coverage.bindings.yaml']}",
        f"  - internals-rest.bindings.yaml: {by_src['internals-rest.bindings.yaml']}",
        f"- Original direction: input {by_dir['input']}, internal {by_dir['internal']}, constant {by_dir['constant']}",
        f"- Remodeled series: **{len(remodeled)}**",
        f"  - semantic / named: {len(named)}",
        f"  - leftover unlifted: {len(unlifted)}",
        "",
        "## Cell coverage",
        "",
        f"- Original unique cells: **{cov['original_cells']}**",
        f"- Remodeled unique cells (data_range minus exclude_rows): **{cov['remodeled_cells']}**",
        f"- Original cells missing from remodeled: **{len(cov['missing'])}**",
        f"- Extra cells introduced by bounding boxes: **{len(cov['extra'])}** (blanks and same-row neighbors are expected)",
        "",
    ]
    if cov["missing"]:
        lines.append("Missing original cells:")
        lines.append("")
        for addr in cov["missing"][:80]:
            lines.append(f"- {addr}")
        if len(cov["missing"]) > 80:
            lines.append(f"- … {len(cov['missing']) - 80} more")
        lines.append("")
    lines += [
        "Extra cells are the bbox fill: group-header rows are excluded via `exclude_rows`,",
        "but blank or other-direction cells on a member row stay inside the rectangle",
        "(schema 1.13.0 has `exclude_rows`, not `exclude_cols`). Input disbursement cells",
        "and internal disbursement cells can therefore share a row and appear as extras",
        "on the other series.",
        "",
        "## Grouping",
        "",
        "Original series were grouped by `(INDICATOR, VARIANT, TABLE, layout, direction)`",
        "and merged when members differ only by `INSTRUMENT` (and `HOLDER` where it is a",
        "single constant on the series). Instrument labels come from original",
        "`series_context.INSTRUMENT` (typos preserved, including `Commecial Bank`) or,",
        "for draft row-blocks, from column B of the member rows via `value_map`.",
        "",
        "Direction is not merged: input vs internal vs constant of the same concept stay",
        "as separate series (`input4_disbursements` vs `input4_disbursements_internal`,",
        "`input4_ida_scale_principal` vs `input4_ida_scale_principal_internal`).",
        "",
        "| Remodeled id | dir | keys | orig series | orig cells | members |",
        "|---|---|---|---:|---:|---|",
    ]
    for s in remodeled:
        members = ", ".join(s["members"][:8])
        if len(s["members"]) > 8:
            members += f", … +{len(s['members']) - 8}"
        keys = ", ".join(s["key"]) if s["key"] else "—"
        lines.append(
            f"| `{s['id']}` | {s['direction']} | {keys} | {s['n_original']} | {s['n_original_cells']} | {members} |"
        )

    lines += ["", "## Reclassified graph-coverage leftovers", ""]
    reclass_notes = [
        "Year-grid series whose `INDICATOR` was an instrument name (`in4_extfin_eurobond_by_year`",
        "and siblings on L:AF of terms rows) were lifted onto `input4_disbursements_internal`",
        "with `INSTRUMENT` from that label and `TIME_PERIOD` from header row 6.",
        "",
        "Column E series keyed by `INDICATOR` via row labels were lifted onto",
        "`input4_discount_rate` (and HOLDER-split FX bond shards) because E is the discount-rate column.",
        "",
        "IDA-scale graph-coverage cells on rows 68–72 (P:BF and the AG:BF scalars) sit on the",
        "principal-repayment schedule, not the terms/disbursement grid. They were lifted onto",
        "`input4_ida_scale_principal_internal` with `TIME_PERIOD` header row **66** (year of loan life).",
        "Several of those leftovers originally bound header row 6 (calendar years on the disbursement",
        "header); that bind is a draft-pass mismatch and is called out below rather than preserved.",
        "",
        "FX local-bond rows reuse instrument labels for non-residents (54:56) and residents (59:61).",
        "`value_map` cannot give one INSTRUMENT key two disjoint row specs, so those blocks are",
        "HOLDER-split series with `bind.kind: constant` for HOLDER (one fixed residency per series).",
        "",
    ]
    lines.extend(reclass_notes)

    if audit.get("header_row_corrections"):
        lines += ["### TIME_PERIOD header_row corrections", ""]
        for item in audit["header_row_corrections"]:
            lines.append(
                f"- `{item['id']}`: original {item['original']} → remodeled `{item['remodeled']}`. {item['reason']}"
            )
        lines.append("")

    if audit.get("dtype_unifications"):
        lines += ["### Measure dtype unifications", ""]
        for item in audit["dtype_unifications"]:
            lines.append(
                f"- `{item['id']}`: original {item['original']} → `{item['remodeled']}` (amounts; int vs float was a draft split)."
            )
        lines.append("")

    if audit.get("divergent_domains"):
        lines += ["### Divergent input domains (omitted on merge)", ""]
        for gid in audit["divergent_domains"]:
            lines.append(f"- `{gid}`")
        lines.append("")
    else:
        lines += [
            "### Input domains",
            "",
            "Grace period members all share `between: {min: 0, max: 50}`; loan maturity members",
            "all share `between: {min: 0, max: 80}`. Those domains are preserved on the merged",
            "input series. Interest and disbursement members had no domain.",
            "",
        ]

    lines += [
        "## Unique one-offs (not merged)",
        "",
        "- `input4_blend_variant` — D16 floating/fixed control for IDA NEW Blend.",
        "- `input4_grace_period_column_header` — G6 column header (original id `in4_extfin_discount_rate`).",
        "- `input4_ida_scale_blend_fixed` — C74 fixed-blend label.",
        "- `input4_ida_regular_translated_name` — D11 translated name.",
        "- `input4_ida_scale_blend_fixed_interest` — D74 `#N/A` placeholder (VARIANT `fixed`).",
        "- `input4_disbursement_totals` — L8:AF8 total disbursements by year (not an instrument row).",
        "",
        "## Leftover unlifted",
        "",
    ]
    if not unlifted:
        lines.append("None.")
        lines.append("")
    else:
        lines.append("| Unlifted id | original id | cells | why |")
        lines.append("|---|---|---:|---|")
        for s in unlifted:
            orig = s["members"][0]
            lines.append(f"| `{s['id']}` | `{orig}` | {s['n_original_cells']} | did not share a merge key after reclassification |")
        lines.append("")

    extra_preview = cov["extra"][:20]
    lines += [
        "## Extra cells",
        "",
        "Bounding-box extras by remodeled series (cells in `data_range` minus `exclude_rows`",
        "that were not in any original Input 4 series). Blanks and same-row neighbors of",
        "sparse principal/disbursement shards dominate.",
        "",
        "| Series | remodeled cells | extras vs original |",
        "|---|---:|---:|",
    ]
    for gid, extra_n, n_cells in cov.get("extra_by_series") or []:
        if extra_n:
            lines.append(f"| `{gid}` | {n_cells} | {extra_n} |")
    lines += [
        "",
        f"Sample of {len(extra_preview)} extra addresses (of {len(cov['extra'])}):",
        "",
    ]
    for addr in extra_preview:
        lines.append(f"- {addr}")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    concept_scheme, originals = load_originals()
    needed_rows = {row for rec in originals for row in rec["rows"]}
    needed_rows |= set(range(8, 76))
    col_b = load_col_b(needed_rows)

    groups: dict[str, list[dict]] = defaultdict(list)
    for rec in originals:
        gid = classify(rec, col_b)
        groups[gid].append(rec)

    audit: dict = {}
    remodeled: list[dict] = []
    ordered_ids = [gid for gid in GROUP_ORDER if gid in groups]
    ordered_ids += sorted(g for g in groups if g not in GROUP_ORDER)

    for gid in ordered_ids:
        members = groups[gid]
        if gid.startswith("input4_unlifted_"):
            if len(members) != 1:
                raise SystemExit(f"unlifted group {gid} has {len(members)} members")
            remodeled.append(build_unlifted(members[0]))
        else:
            remodeled.append(build_group_series(gid, members, col_b, audit))

    ids = [s["id"] for s in remodeled]
    if len(ids) != len(set(ids)):
        dup = [i for i in ids if ids.count(i) > 1]
        raise SystemExit(f"duplicate ids: {sorted(set(dup))}")

    cov = coverage(originals, remodeled)
    if cov["missing"]:
        print(f"WARNING: {len(cov['missing'])} original cells not covered")
        print(cov["missing"][:30])

    header = (
        'schema_version: "1.13.0"\n'
        "workbook: workbook.xlsm\n"
        f"{emit_concept_scheme(concept_scheme)}\n"
        "series:\n"
    )
    body = "\n".join(emit_series(s) for s in remodeled) + "\n"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "input4.bindings.yaml").write_text(header + body)
    (OUT_DIR / "input4-audit.md").write_text(
        write_audit(originals, groups, remodeled, cov, audit)
    )
    print(f"original series: {len(originals)}")
    print(f"remodeled series: {len(remodeled)}")
    print(f"unlifted: {sum(1 for s in remodeled if s['id'].startswith('input4_unlifted_'))}")
    print(f"original cells: {cov['original_cells']}")
    print(f"remodeled cells: {cov['remodeled_cells']}")
    print(f"missing: {len(cov['missing'])} extra: {len(cov['extra'])}")
    print("top series:")
    for s in remodeled[:12]:
        print(f"  {s['id']} keys={s['key']} n_orig={s['n_original']}")


if __name__ == "__main__":
    main()
